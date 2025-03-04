'''
		A collection of useful Methods	
'''
import os
import glob
import KapchaRosskyScale
import ModelParameters
import Tools

from matplotlib.pyplot import rc
rc('font', **{'family': 'serif', 'serif': ['Computer Modern'], 'size':45})
rc('text', usetex=True)

import numpy as np
import matplotlib.pyplot as plt

import math
from scipy import stats
import openpyxl
import scipy.optimize


PDB = os.getcwd()[-4:]
'''
		Analyze HOLE output
'''
def HOLEOutputAnalysis():

	mol_info_fn = glob.glob('*_molInfo.xlsx')[0]
	wb = openpyxl.load_workbook(mol_info_fn)	 
	source = wb['geometry_information']
	ind_ppAxis	= source.cell(row = 2, column = 1).value 
	axis_ori = np.sign(ind_ppAxis) 
	ind_ppAxis = abs(ind_ppAxis) - 1
	
	## Load data 
	geomCenter = np.loadtxt(glob.glob('*_gc.dat')[0])
	
	## Initialize ..
	mins = np.empty((ModelParameters.N_RUNS_HOLE, 3))
	maxs = np.empty((ModelParameters.N_RUNS_HOLE, 3))
	mean_diffs = np.empty((ModelParameters.N_RUNS_HOLE, 3))
	std_diffs = np.empty((ModelParameters.N_RUNS_HOLE, 3))
	porePoints_x = []
	porePoints_y = []
	porePoints_z = []
	poreRadii = []
	
	'''
			Step 2. Collect the data
	'''
	subdir = PDB + "_prad" 
	for i in range(ModelParameters.N_RUNS_HOLE):
			
		## Load data
		porePoints = np.loadtxt(f'{subdir}/pp_{i}.dat')
		poreRadius = np.loadtxt(f'{subdir}/ppr_{i}.dat')
		
		## Introduce global origin at geomCenter
		porePoints = porePoints - geomCenter
		poreRadius[:,0] = poreRadius[:,0] - geomCenter[ind_ppAxis]
		
		## Make sure you get all the pore points (and corresponding pore radii)
		porePoints_ordered = np.empty((len(poreRadius), 3))
		for j in range(len(poreRadius)):
			p_point = poreRadius[j,0]
			p_coords = porePoints[:,ind_ppAxis] 
			abs_diff = abs(p_coords - p_point)
			ind = np.where(min(abs_diff) == abs_diff)
			if (len(ind) != 1):												# there should be only one match .. 
				exit('\n .. Exiting smoothly .. No match .. \n')
			else:
				porePoints_ordered[j,:] = np.round(porePoints[ind,:], ModelParameters.ROUND_DIGIT_PP) # round
				
		
		porePoints = porePoints_ordered
		poreRad = poreRadius[:,1]
		
		# mins	
		mins[i,0] = min(porePoints[:,0])
		mins[i,1] = min(porePoints[:,1])
		mins[i,2] = min(porePoints[:,2])
		# maxs
		maxs[i,0] = max(porePoints[:,0])
		maxs[i,1] = max(porePoints[:,1])
		maxs[i,2] = max(porePoints[:,2])
		# std_diffs diffs
		std_diffs[i,0] = np.std(np.diff(porePoints[:,0]))	
		std_diffs[i,1] = np.std(np.diff(porePoints[:,1]))	
		std_diffs[i,2] = np.std(np.diff(porePoints[:,2]))	
		# means diffs
		mean_diffs[i,0] = np.mean(np.diff(porePoints[:,0]))	
		mean_diffs[i,1]	= np.mean(np.diff(porePoints[:,1]))	
		mean_diffs[i,2]	= np.mean(np.diff(porePoints[:,2]))	
		
		# plt.plot(porePoints[:,2], poreRad)
		# plt.ylabel(r'R(\textbf{p}) [\AA]', fontsize = 26)
		# plt.xlabel(r'$\perp$-coord.', fontsize = 26)
		# plt.tight_layout()
		# plt.show()
		
		poreRadii.append(poreRad)
		porePoints_x.append(porePoints_ordered[:,0])
		porePoints_y.append(porePoints_ordered[:,1])
		porePoints_z.append(porePoints_ordered[:,2])
	
	'''
			Step 3. Create a common axis
	'''
	if (np.mean(std_diffs[:,2]) < ModelParameters.ACC): 
		
		dz_round = np.round(np.mean(mean_diffs[:,2]), ModelParameters.ROUND_DIGIT_PP)
		l = np.round(max(mins[:,2]), ModelParameters.ROUND_DIGIT_PP)
		r = np.round(min(maxs[:,2]), ModelParameters.ROUND_DIGIT_PP)
		common_axis = np.arange(l, r + dz_round, dz_round)
		porePoints_1_trimmed = np.empty((ModelParameters.N_RUNS_HOLE, len(common_axis)))
		porePoints_2_trimmed = np.empty((ModelParameters.N_RUNS_HOLE, len(common_axis)))
		poreRadii_trimmed = np.empty((ModelParameters.N_RUNS_HOLE, len(common_axis)))
		print('\n Channel should be aligned with z-axis (i.e., ind_ppAxis is (%d = 2)) ' % ind_ppAxis)
		if (ind_ppAxis != 2):
			exit('\n\n .. Exiting smoothly .. The provided alignment information is incorrect! \n\n')
		##  Calculate the mean_diffs and std_diffs of xy coords
		for i in range(ModelParameters.N_RUNS_HOLE): 
			for j in range(len(common_axis)):
				ind = np.where(abs(common_axis[j] - np.round(porePoints_z[i], ModelParameters.ROUND_DIGIT_PP)) < dz_round + ModelParameters.ACC)[0][0]
				##  Keep what is really common (intersection)
				porePoints_1_trimmed[i,j] = (porePoints_x[i])[ind]
				porePoints_2_trimmed[i,j] = (porePoints_y[i])[ind]
				poreRadii_trimmed[i,j] = (poreRadii[i])[ind]
		
	else:
		if (np.mean(std_diffs[:,0]) < ModelParameters.ACC):
			
			dx_round = np.round(np.mean(mean_diffs[:,0]), ModelParameters.ROUND_DIGIT_PP)
			l = np.round(max(mins[:,0]), ModelParameters.ROUND_DIGIT_PP)
			r = np.round(min(maxs[:,0]), ModelParameters.ROUND_DIGIT_PP)
			common_axis = np.arange(l, r + dx_round, dx_round)
			porePoints_1_trimmed = np.empty((ModelParameters.N_RUNS_HOLE, len(common_axis)))
			porePoints_2_trimmed = np.empty((ModelParameters.N_RUNS_HOLE, len(common_axis)))
			poreRadii_trimmed = np.empty((ModelParameters.N_RUNS_HOLE, len(common_axis)))
			print('\n Channel should be aligned with x-axis (i.e., ind_ppAxis is (%d = 0)) ' % ind_ppAxis)
			if (ind_ppAxis != 0):
				exit('\n\n .. Exiting smoothly .. The provided alignment information is incorrect! \n\n')
			##	Calculate the mean_diffs and std_diffs of yz coords
			for i in range(ModelParameters.N_RUNS_HOLE): 
				for j in range(len(common_axis)):
					ind = np.where(abs(common_axis[j] - np.round(porePoints_x[i], ModelParameters.ROUND_DIGIT_PP)) < dx_round + ModelParameters.ACC)[0][0]
					## 	Keep what is really common (intersection)
					porePoints_1_trimmed[i,j] = (porePoints_y[i])[ind]
					porePoints_2_trimmed[i,j] = (porePoints_z[i])[ind]
					poreRadii_trimmed[i,j] = (poreRadii[i])[ind]
		
		if (np.mean(std_diffs[:,1]) < ModelParameters.ACC): 
			
			dy_round = np.round(np.mean(mean_diffs[:,1]), ModelParameters.ROUND_DIGIT_PP)
			l = np.round(max(mins[:,1]), ModelParameters.ROUND_DIGIT_PP)
			r = np.round(min(maxs[:,1]), ModelParameters.ROUND_DIGIT_PP)
			common_axis = np.arange(l, r + dy_round, dy_round)
			porePoints_1_trimmed = np.empty((ModelParameters.N_RUNS_HOLE, len(common_axis)))
			porePoints_2_trimmed = np.empty((ModelParameters.N_RUNS_HOLE, len(common_axis)))
			poreRadii_trimmed = np.empty((ModelParameters.N_RUNS_HOLE, len(common_axis)))
			print('\n Channel should be aligned with y-axis (i.e., ind_ppAxis is (%d = 1)) ' % ind_ppAxis)
			if (ind_ppAxis != 1):
				exit('\n\n .. Exiting smoothly .. The provided alignment information is incorrect! \n\n')
			##  Calculate the mean_diffs and std_diffs of xz coords
			for i in range(ModelParameters.N_RUNS_HOLE): 
				for j in range(len(common_axis)):
					ind = np.where(abs(common_axis[j] - np.round(porePoints_y[i], ModelParameters.ROUND_DIGIT_PP)) < dy_round + ModelParameters.ACC)[0][0]
					##  Keep what is really common (intersection)
					porePoints_1_trimmed[i,j] = (porePoints_x[i])[ind]
					porePoints_2_trimmed[i,j] = (porePoints_z[i])[ind]
					poreRadii_trimmed[i,j] = (poreRadii[i])[ind]
		
	'''
			4. Calculate some useful stats
	'''
	mean_poreRad = np.zeros(len(common_axis))
	mean_coord1	= np.zeros(len(common_axis))
	mean_coord2 = np.zeros(len(common_axis))
	std_poreRad	= np.zeros(len(common_axis))
	
	for j in range(len(common_axis)):
		mean_poreRad[j] = np.mean(poreRadii_trimmed[:,j])
		mean_coord1[j] = np.mean(porePoints_1_trimmed[:,j])
		mean_coord2[j] = np.mean(porePoints_2_trimmed[:,j])
		if (j == 50):
			plt.plot(porePoints_1_trimmed[:,j], porePoints_2_trimmed[:,j], 'o')
			plt.plot(mean_coord1[j], mean_coord2[j], 'x')
			plt.tight_layout()
			plt.savefig(PDB + '_hole_plane.pdf')
			plt.clf()
			plt.close()
		std_poreRad[j] = np.std(poreRadii_trimmed[:,j])
	
	'''
			5. Make a figure and store the data
	'''
	plt.plot(axis_ori*common_axis, mean_poreRad, 'b', alpha=0.75, linewidth = 2.0, label = r'$R$') # nu
	plt.fill_between(axis_ori*common_axis, mean_poreRad - std_poreRad, mean_poreRad + std_poreRad, color='b', alpha=0.3)
	plt.ylabel(r'R(\textbf{p}) [\AA]', fontsize = 26)
	plt.xlabel(r'$\perp$-coord.', fontsize = 26)
	plt.tight_layout()
	plt.savefig(PDB + '_poreRadius_mean.pdf')
	plt.clf()
	plt.close()
	
	## store the mean (average) pore radius## store the mean (average) pore radius
	np.savetxt(PDB + '_mean_poreRad.dat', np.column_stack((mean_poreRad, std_poreRad)))
	
	## store the pore points coordinates
	if (ind_ppAxis == 2):
		porePoints = np.column_stack((mean_coord1, mean_coord2, common_axis))
		np.savetxt(PDB + '_mean_porePoints.dat', porePoints)
	if (ind_ppAxis == 1):
		porePoints = np.column_stack((mean_coord1, common_axis, mean_coord2))
		np.savetxt(PDB + '_mean_porePoints.dat', porePoints)	
	if (ind_ppAxis == 0):
		porePoints = np.column_stack((common_axis, mean_coord1, mean_coord2))
		np.savetxt(PDB + '_mean_porePoints.dat', porePoints)
	
	plt.show()


'''
		PDB structure preparation
'''
def PDBStructurePreperation():

	# Import locally ..
	from itertools import permutations
	from Bio.PDB import PDBParser

	'''
			Load HOLE data 	
	'''
	# Find file names ..
	pdb_fn = glob.glob('*_H_ori_renum.pdb')[0]					# renumbered! no proper numbering	
	poreRadius_fn = glob.glob('*_mean_poreRad.dat')[0]
	porePoints_fn = glob.glob('*_mean_porePoints.dat')[0]

	# Load file content ..
	poreRadius = np.loadtxt(poreRadius_fn)
	porePoints = np.loadtxt(porePoints_fn)
	poreRadius_mean = poreRadius[:,0]	
	poreRadius_std = poreRadius[:,1]	

	'''
			PDB structure checking and collection of relevant mol chars
	'''
	f = open('Preparation_Report.txt', 'a')
	
	# Parse and get basic information
	parser = PDBParser()
	struct = parser.get_structure('STS', pdb_fn)
	
	# Initialize ..
	totalNrOfAtoms = 0
	flag_disordered = 0
	flag_het = 0
	flag_missmatch = 0
	flag_missing = 0
	atomID = []
	bfactors = []
	vdWrad = []
	hvals = []
	occupancy = []
	coords_x = []
	coords_y = []
	coords_z = []
	resInfo	= []
	coords_x_res = []
	coords_y_res = []
	coords_z_res = []
	chains = []

	'''	
			Step 1. Check structure! Assign to each atom a hydropathic value. Collect all relevant information and perform statistical analysis (distr. fitting).
	'''
	for model in struct:
		for chain in model:
			chains.append([chain.id, len(chain)])
			for residue in chain:	
				if (residue.is_disordered() == 1): flag_disordered += 1
				if (residue.get_id()[0] == 1): flag_het += 1
				res_name = residue.get_resname()
				res_hydro = 0
				res_nrOfAtoms = 0
				res_x = 0
				res_y = 0
				res_z = 0
				for atom in residue:
					atom_name = atom.get_name()
					# atom.set_bfactor()  
					atom_x,atom_y,atom_z = atom.get_coord()
					hydroChar = KapchaRosskyScale.df[res_name].iloc[:].dropna()
					expected_nrOfAtoms = len(hydroChar)
					hydroChar_ind = hydroChar.index[:]
					hydroChar_val = hydroChar.loc[:]
					perms = [''.join(p) for p in permutations(atom_name)] 	# this makes sure that HD12 and 1HD2 are indeed the same thing
					for i in range(len(perms)):
						ind_atom = np.where(hydroChar_ind == perms[i])[0]
						if (len(ind_atom) != 0):
							break
					noise = np.random.normal(0, ModelParameters.STD_NOISE, 1)
					if (len(ind_atom) == 0): 
						print('\n\n Warning! Atom could not be mapped on the KR scale: ', atom_name, res_name, hydroChar_val[ind_atom], file = f)
						# print('\n\n Warning! Atom (%s,%s,%f) could not be mapped on the KR scale \n\n' % (atom_name, res_name, hydroChar_val[ind_atom]), file = f)
						hvals = np.append(hvals, noise)
						print('\n Orfan (unmapped) atom was assigned a hydropathic value of %f \n' % noise, file = f)
						flag_missmatch += 1
					else:
						res_hydro += hydroChar_val[ind_atom[0]] + noise
						hvals = np.append(hvals, hydroChar_val[ind_atom[0]] + noise)
					bfactors = np.append(bfactors, atom.get_bfactor())
					occupancy = np.append(occupancy, atom.get_occupancy())
					atomID = np.append(atomID, atom_name)
					coords_x = np.append(coords_x, atom_x)
					coords_y = np.append(coords_y, atom_y)
					coords_z = np.append(coords_z, atom_z)
					vdWrad_val = Tools.VdWRad(atom_name)
					if (vdWrad_val == None):
						exit('\n\n .. Exiting smoothly .. Check your vdW radii assignments .. \n\n')
					else:
						vdWrad = np.append(vdWrad, vdWrad_val)	
					totalNrOfAtoms += 1
					res_x += atom_x
					res_y += atom_y
					res_z += atom_z
					res_nrOfAtoms += 1
				if (res_nrOfAtoms != expected_nrOfAtoms):
					print('\n\n Warning! Missing %d atoms (out of %d) from residue %s %s' % (expected_nrOfAtoms - res_nrOfAtoms , expected_nrOfAtoms, residue.get_resname(), residue.get_id()[1]), file=f)
					flag_missing += 1
				resInfo.append([residue.get_id()[1], residue.get_resname(), res_hydro])
				# print([residue.get_id()[1], residue.get_resname(), res_hydro])
				coords_x_res = np.append(coords_x_res, res_x/res_nrOfAtoms )
				coords_y_res = np.append(coords_y_res, res_y/res_nrOfAtoms )
				coords_z_res = np.append(coords_z_res, res_z/res_nrOfAtoms )
	
	
	# Get the geom center and print some useful information			
	geomCenter = [sum(coords_x)/totalNrOfAtoms , sum(coords_y)/totalNrOfAtoms , sum(coords_z)/totalNrOfAtoms ]
	print('\n The geom center of the molecule is: %s' % str(geomCenter)[1:-1], file=f)
	print(' There were %d missmatches found.' % flag_missmatch, file=f)
	print(' There were %d misssing atoms.' % flag_missing, file=f)
	print(' There were %d heterogeneous atoms found.' % flag_het, file=f)
	print(' There were %d disordered atoms found.' % flag_disordered, file=f)
	print(' There were %d atoms in total.' % totalNrOfAtoms , file=f)
	print(' There were %d residues in total! \n' % len(resInfo), file=f)

	'''
			Step 2. Prepare coordinates
	'''
	# Stack some useful information ..
	coords_atom = np.column_stack((coords_x, coords_y, coords_z))
	coords_res = np.column_stack((coords_x_res, coords_y_res, coords_z_res))
	# Introduce global origin at geomCenter
	coords_atom = coords_atom - geomCenter
	coords_res = coords_res - geomCenter
	# Plot the top view so that you can visually easily check if the molecule is properly aligned
	plt.plot(coords_atom[:,0], coords_atom[:,1], 'bo', alpha = 0.15, markersize = 1)
	plt.plot(coords_res[:,0], coords_res[:,1], 'rx', alpha = 0.5, markersize = 2)
	plt.savefig(PDB + '_topView.pdf')
	plt.clf()
	plt.close()

	'''
			Step 3. Store useful molecular information
	'''
	## Store some useful PDB structure information
	mol = {}
	mol['porePoints'] = porePoints
	mol['poreRadius_mean'] = poreRadius_mean
	mol['poreRadius_std'] = poreRadius_std
	mol['coords_atom'] = coords_atom
	mol['coords_res'] = coords_res
	mol['hvals'] = hvals
	mol['resInfo'] = resInfo
	mol['vdWrad'] = vdWrad
	mol['total_nrOfAtoms'] = totalNrOfAtoms 
	mol['chains'] = chains

	Tools.StoreFile(mol, 'mol')

'''
		Insert Variants (hotspot structural locations)
'''
def InsertVariants():
	
	# Import locally ..
	from collections import Counter
	from Bio.PDB import PDBParser

	mol = Tools.LoadFile('mol')
	
	# We need the properly numbered pdb file to assign the variants locations
	pdb_fn = glob.glob('*_H_ori.pdb')[0]								
	
	# Process pain-causing variants information 
	fn = glob.glob('*_molInfo.xlsx')[0]
	wb = openpyxl.load_workbook(fn)	 
	source = wb['variants_information']
	info_marker = source.cell(row = 2, column = 1).value 	
	
	if (info_marker == 0):
		print('\n\n .. No variants are selected and/or inserted ..')
		return
	
	# Process gnomad information 
	fn = glob.glob('*data_gnomaD*')[0]
	gnomadData = openpyxl.load_workbook(fn)	 
	source_g = gnomadData['gnomad_information']
	
	# Process clinvar information 
	fn = glob.glob('*data_ClinVar*')[0]
	clinvarData = openpyxl.load_workbook(fn)	 
	source_c = clinvarData['clinVar_information']

	if (info_marker == 1):	
			
		offset = source.cell(row = 4, column = 1).value 
		offsetID = source.cell(row = 6, column = 1).value 
		
		'''
				IEM
		'''			
		iem_id = np.array(Tools.VariantsListing(source['B']))
		iem_resType = Tools.VariantsListing(source['C'])
		# Introduce offset!
		inds = np.where(iem_id > offsetID)[0]
		# normalize: if (list_id > offsetID) then pdb_id = list_id + offset
		iem_id[inds] = iem_id[inds.astype(int)] + offset
		
		'''
				SFN
		'''		
		sfn_id = np.array(Tools.VariantsListing(source['D']))
		sfn_resType = Tools.VariantsListing(source['E'])
		# Introduce offset!
		inds = np.where(sfn_id > offsetID)[0]
		# normalize: if (list_id > offsetID) then pdb_id = list_id + offset
		sfn_id[inds] = sfn_id[inds.astype(int)] + offset
		
		'''
				PEPD
		'''
		pepd_id = np.array(Tools.VariantsListing(source['F']))
		pepd_resType = Tools.VariantsListing(source['G'])
		# Introduce offset!
		inds = np.where(pepd_id > offsetID)[0]
		# normalize: if (list_id > offsetID) then pdb_id = list_id + offset
		pepd_id[inds] = pepd_id[inds.astype(int)] + offset
		
		'''
				NEUTRAL
		'''
		neutral_id = np.array(Tools.VariantsListing(source['H']))
		neutral_resType = Tools.VariantsListing(source['I'])
		# Introduce offset!
		inds = np.where(neutral_id > offsetID)[0]
		# normalize: if (list_id > offsetID) then pdb_id = list_id + offset 
		neutral_id[inds] = neutral_id[inds.astype(int)] + offset
		
		'''
				LOF
		'''
		lof_id = np.array(Tools.VariantsListing(source['J']))
		lof_resType = Tools.VariantsListing(source['K'])
		# Introduce offset!
		inds = np.where(lof_id > offsetID)[0]
		# normalize: if (list_id > offsetID) then pdb_id = list_id + offset
		lof_id[inds] = lof_id[inds.astype(int)] + offset

		'''
				MISSCLASS
		'''
		miss_id = np.array(Tools.VariantsListing(source['M']))
		miss_resType = Tools.VariantsListing(source['N'])
		# Introduce offset!
		inds = np.where(miss_id > offsetID)[0]
		# normalize: if (list_id > offsetID) then pdb_id = list_id + offset
		miss_id[inds] = miss_id[inds.astype(int)] + offset
						
		'''		
				GNOMAD (benign, pathognic, vus)
		'''
		gnomad_id = np.array(Tools.VariantsListing(source_g['A']))
		element_counts = Counter(gnomad_id)
		# Filter elements that appear more than once
		duplicates = [item for item, count in element_counts.items() if count > 1]
		if (len(duplicates) > 0):
			print('.. There are list entries that appear more than once (gnomad): ', duplicates)
		gnomad_resType = Tools.VariantsListing(source_g['B'])
		gnomad_significance = Tools.VariantsListing(source_g['D']) 
		gnomad_indices = Tools.VariantsListing(source_g['E']) 
		# note that: NO offset needed!
		
		'''		
				ClinVar (missence variants)
		'''
		clinvar_id = np.array(Tools.VariantsListing(source_c['A']))
		element_counts = Counter(clinvar_id)
		# Filter elements that appear more than once
		duplicates = [item for item, count in element_counts.items() if count > 1]
		if (len(duplicates) > 0):
			print('.. There are list entries that appear more than once (ClinVar): ', duplicates)
		clinvar_resType = Tools.VariantsListing(source_c['B'])
		clinvar_significance = Tools.VariantsListing(source_c['D']) 
		clinvar_indices = Tools.VariantsListing(source_c['E']) 
		# note that: NO offset needed!
		
		print('\n Note that there are %d ClinVar entries not found in Gnomad!' % len(set(clinvar_id) - set(gnomad_id)))
		# print('They are: ', set(clinvar_id) - set(gnomad_id))
		
		# Parse and get basic information
		parser = PDBParser()
		struct = parser.get_structure('STS', pdb_fn)
		
		# Initialize ..
		n_iem = 0
		n_pepd = 0
		n_sfn = 0
		n_neutral = 0
		n_lof = 0
		n_miss = 0
		n_gnomad = 0
		n_clinvar = 0
		# nine categories from Gnomad!
		n_unknown_gnomad = 0
		n_benign_gnomad = 0
		n_likelyBenign_gnomad = 0
		n_benignLikelyBenign_gnomad = 0	
		n_pathogenic_gnomad	= 0
		n_likelyPathogenic_gnomad = 0
		n_pathogenicLikelyPathogenic_gnomad	= 0
		n_conflInter_gnomad = 0
		n_vus_gnomad = 0
		# ClinVar
		n_vus_clinvar = 0
		n_unknown_clinvar = 0
		n_benign_clinvar = 0
		n_likelyBenign_clinvar = 0
		n_benignLikelyBenign_clinvar = 0	
		n_pathogenic_clinvar = 0
		n_likelyPathogenic_clinvar = 0
		n_pathogenicLikelyPathogenic_clinvar = 0
		n_conflInter_clinvar = 0
		n_vus_clinvar = 0
		
		nrOfRes = len(mol['resInfo'])
		# Find out residue IDs of domain I, II, III, and IV
		nrOfRes_chain = 0
		# Find out the id of the longest chain in the model (should be the alpha subunit)
		for i in range(len(mol['chains'])):
			if (mol['chains'][i][1] > nrOfRes_chain):
				nrOfRes_chain = mol['chains'][i][1]
				chain_id = i
				
		# ...		
		iem_list_retrieved = []		
		pepd_list_retrieved = []		
		sfn_list_retrieved = []		
		lof_list_retrieved = []
		miss_list_retrieved = []
		neutral_list_retrieved = []
		gnomad_list_retrieved = [] # gnomad
		clinvar_list_retrieved = [] # clinvar
		
		# ...
		# All the information that we need: 
		# First column: residue id (according to the Pain Information nomeclature!) 
		# Second column: variant characterization
		# Third column: variant (residue) name
		# Fourth column: variant classification status
		varInfo = [[None for _ in range(4)] for _ in range(nrOfRes)]

		# ...
		i = 0	# all residues
		j = 0 	# residues only in Domain I, II, III, and IV
		nrOfDomains = 1
		domainBoundaries = {}
		for model in struct:
			for chain in model:
				for residue in chain:	
					
					res_id = residue.get_id()[1]
					res_name = residue.get_resname()
				
					varInfo[i][0] = int(res_id)
					varInfo[i][1] = 'not provided'
					varInfo[i][2] = res_name
					varInfo[i][3] = 'classified'
										
					## Count how many residues there in the alpha subunit and 
					## retrieve domain I, II, III, and IV residues indices (i index!)	
					if (chain.id == mol['chains'][chain_id][0]): 
						j += 1
						if (j == 1):
							domainBoundaries['Domain_I_ResID'] = res_id
							domainBoundaries['Domain_I_Index'] = i
							# print(res_id, i)
						if (j > 1):
							if (res_id - res_id_pre > 1):
								nrOfDomains += 1 
								if (nrOfDomains == 2):
									domainBoundaries['Domain_II_ResID'] = res_id
									domainBoundaries['Domain_II_Index'] = i
									# print(res_id, i)	
								if (nrOfDomains == 3):
									domainBoundaries['Domain_III_ResID'] = res_id
									domainBoundaries['Domain_III_Index'] = i
									# print(res_id, i)	
								if (nrOfDomains == 3):
									domainBoundaries['Domain_IV_ResID'] = res_id
									domainBoundaries['Domain_IV_Index'] = i
									# print(res_id, i)	
						res_id_pre = res_id


					"""

							Check for variant missclassification (this will concerns only the fourth column of varInfo)
					
					"""
					if (res_id in miss_id and chain.id == mol['chains'][chain_id][0]):
						
						ind = Tools.Match(res_id, miss_id)
						
						if (len(ind) < 2):
							if (res_name == miss_resType[ind[0]]):
								
								# Add the event to the corresponding list
								miss_list_retrieved.append(res_id)								
								# Change the fourth column of varInfo: 
								varInfo[i][3] = 'missclassified'
								# Count
								n_miss += 1
							
							else:
								exit('\n\n .. Found id, but type is not matching! (miss) .. \n\n')
						else:
							exit('\n\n .. Too many matches .. (miss)   \n\n')


					"""
							Check for variant phenotype:
							- IEM
							- SFN
							- PEPD
							- LoF
							- Neutrals (gnomad, ClinVar)
					"""
					## Insert IEM mutations: index 1 ##
					if (res_id in iem_id and chain.id == mol['chains'][chain_id][0]):	
						
						ind = Tools.Match(res_id, iem_id)
							
						if (len(ind) < 2):	
							if (res_name == iem_resType[ind[0]]):
								
								# Add the event to the corresponding list
								iem_list_retrieved.append(res_id) 
								# Change the second column of varInfo: 
								varInfo[i][1] = 'IEM'
								# Count
								n_iem += 1
								# Increment!
								i += 1 
								# Continue!
								continue

							else:
								exit('\n\n .. Found id, but type is not matching! (IEM) .. \n\n')
						else:
							exit('\n\n .. Too many matches .. (IEM)  \n\n')
					
					## Insert SFN mutations: index 2 ##
					if (res_id in sfn_id and chain.id == mol['chains'][chain_id][0]):
						
						ind = Tools.Match(res_id, sfn_id)
						
						if (len(ind) < 2):
							if (res_name == sfn_resType[ind[0]]):
								
								# Add the event to the corresponding list
								sfn_list_retrieved.append(res_id)
								# Change the second column of varInfo: 
								varInfo[i][1] = 'SFN'
								# Count
								n_sfn += 1
								# Increment!
								i += 1 
								# Continue!
								continue

							else:
								exit('\n\n .. Found id, but type is not matching! (SFN) .. \n\n')
						else:
							exit('\n\n .. Too many matches .. (SFN)   \n\n')	
					
					## Insert PEPD mutations: index 3 ##
					if (res_id in pepd_id and chain.id == mol['chains'][chain_id][0]):

						ind = Tools.Match(res_id, pepd_id)
						
						if (len(ind) < 2):
							if (res_name == pepd_resType[ind[0]]):
								
								# Add the event to the corresponding list
								pepd_list_retrieved.append(res_id)
								# Change the second column of varInfo: 
								varInfo[i][1] = 'PEPD'
								# Count
								n_pepd += 1
								# Increment!
								i += 1 
								# Continue!
								continue

							else:
								exit('\n\n .. Found id, but type is not matching! (PEPD) .. \n\n')
						else:
							exit('\n\n .. Too many matches .. (PEPD)   \n\n')

					## Insert NEUTRAL mutations: index 4 ##						
					if (res_id in neutral_id and chain.id == mol['chains'][chain_id][0]):
						
						ind = Tools.Match(res_id, neutral_id)							
						
						if (len(ind) < 2):
							if (res_name == neutral_resType[ind[0]]):
				
								# Add the event to the corresponding list
								neutral_list_retrieved.append(res_id)
								# Change the second column of varInfo: 
								varInfo[i][1] = 'Neutral'
								# Count
								n_neutral += 1
								# Increment!
								i += 1 
								# Continue!
								continue

							else:
								exit('\n\n .. Found id, but type is not matching! (Neutral) .. \n\n')
						else:
							exit('\n\n .. Too many matches .. (Neutral)   \n\n')
							
					
					
					## Insert lof mutations: index 5 ##
					if (res_id in lof_id and chain.id == mol['chains'][chain_id][0]):
						
						ind = Tools.Match(res_id, lof_id)
						
						if (len(ind) < 2):
							if (res_name == lof_resType[ind[0]]):
								
								# Add the event to the corresponding list
								lof_list_retrieved.append(res_id)
								# Change the second column of varInfo: 
								varInfo[i][1] = 'LoF'
								# Count
								n_lof += 1
								# Increment!
								i += 1 
								# Continue!
								continue
							
							else:
								exit('\n\n .. Found id, but type is not matching! (LoF) .. \n\n')
						else:
							exit('\n\n .. Too many matches .. (LoF)   \n\n')
					
						
					## Insert gnomad mutations! several indices depending on characterization
					if (res_id in gnomad_id and chain.id == mol['chains'][chain_id][0]):
						
						# print(res_id)
						ind = Tools.Match(res_id, gnomad_id)
										
						if (len(ind) < 2):
							if (res_name != gnomad_resType[ind[0]]):		
							
								print(' (Gnomad non-hit) Residue id is NOT found: %d (%s, %s), (%f,%f,%f)' % (res_id, res_name, gnomad_resType[ind[0]], mol['coords_res'][i][0], mol['coords_res'][i][1], mol['coords_res'][i][2]))
							
							# Add the event to the corresponding list
							gnomad_list_retrieved.append(res_id)

							# Change the second column of varInfo: 
							varInfo[i][1] = gnomad_significance[ind[0]]

							# Categorize:
							if (gnomad_indices[ind[0]] == 0):
								n_unknown_gnomad += 1
								
							if (gnomad_indices[ind[0]] == -1):
								n_benign_gnomad += 1
								
							if (gnomad_indices[ind[0]] == -2):
								n_benignLikelyBenign_gnomad += 1
								
							if (gnomad_indices[ind[0]] == -3):
								n_likelyBenign_gnomad += 1
								
							if (gnomad_indices[ind[0]] == -4):
								n_pathogenic_gnomad += 1
								
							if (gnomad_indices[ind[0]] == -5):
								n_pathogenicLikelyPathogenic_gnomad += 1
								
							if (gnomad_indices[ind[0]] == -6):
								n_likelyPathogenic_gnomad += 1
								
							if (gnomad_indices[ind[0]] == -7):
								n_conflInter_gnomad += 1
								
							if (gnomad_indices[ind[0]] == -8):
								n_vus_gnomad += 1
							
							# Count
							n_gnomad += 1
							# Increment!
							i += 1 
							# Continue!
							continue
								
						else:
							exit('\n\n .. Too many matches .. (GNOMAD vars)   \n\n')
							
													
							
					## Insert clinvar mutations! several indices depending on characterization
					if (res_id in clinvar_id and chain.id == mol['chains'][chain_id][0]):
						
						
						ind = Tools.Match(res_id, clinvar_id)
										
						if (len(ind) < 2):
							if (res_name != clinvar_resType[ind[0]]):		
								
								print(' (ClinVar non-hit) Residue id is NOT found: %d, (%s = %s), (%f,%f,%f)' % (res_id, res_name, clinvar_resType[ind], mol['coords_res'][i][0], mol['coords_res'][i][1], mol['coords_res'][i][2]))

							# Add the event to the corresponding list	
							clinvar_list_retrieved.append(res_id)
								
							# Change the second column of varInfo: 
							varInfo[i][1] = clinvar_significance[ind[0]]
								
							# Categorize:
							if (clinvar_indices[ind[0]] == 0):
								n_unknown_clinvar += 1
								
							if (clinvar_indices[ind[0]] == -1):
								n_benign_clinvar += 1
								
							if (clinvar_indices[ind[0]] == -2):
								n_benignLikelyBenign_clinvar += 1
								
							if (clinvar_indices[ind[0]] == -3):
								n_likelyBenign_clinvar += 1
								
							if (clinvar_indices[ind[0]] == -4):
								n_pathogenic_clinvar += 1
								
							if (clinvar_indices[ind[0]] == -5):
								n_pathogenicLikelyPathogenic_clinvar += 1
								
							if (clinvar_indices[ind[0]] == -6):
								n_likelyPathogenic_clinvar += 1
								
							if (clinvar_indices[ind[0]] == -7):
								n_conflInter_clinvar += 1
								
							if (clinvar_indices[ind[0]] == -8):
								n_vus_clinvar += 1

							# Count	
							n_clinvar += 1
							# Increment!
							i += 1 
							# Continue!
							continue
									
						else:
							exit('\n\n .. Too many matches .. (ClinVar vars)   \n\n')
							
						
					# Increment!				
					i += 1
		
		if (n_iem < len(iem_id)):			
			print('\n We are missing %d iem variant(s) with id(s): ' % (len(iem_id) - n_iem))
			print(' %s ' % list(set(iem_list_retrieved) ^ set(iem_id)))	
		elif (n_iem == len(iem_id)): 
			if (set(iem_list_retrieved) != set(iem_id)):
				print(iem_list_retrieved)
				print(iem_id)
				exit('\n\n ... Exiting smoothly ... Missmatch between variants sets (IEM) \n\n')
			else:
				print(' .. All IEM entries were collected! .. ')
			
		if (n_sfn < len(sfn_id)):			
			print('\n We are missing %d sfn variant(s) with id(s): ' % (len(sfn_id) - n_sfn))
			print(' %s ' % list(set(sfn_list_retrieved) ^ set(sfn_id)))	
		elif (n_sfn == len(sfn_id)): 
			if (set(sfn_list_retrieved) != set(sfn_id)):
				print(sfn_list_retrieved)
				print(sfn_id)
				exit('\n\n ... Exiting smoothly ... Missmatch between variants sets (SFN) \n\n')
			else:
				print(' .. All SFN entries were collected! .. ')
		
		if (n_pepd < len(pepd_id)):			
			print('\n We are missing %d pepd variant(s) with id(s): ' % (len(pepd_id) - n_pepd))
			print(' %s ' % list(set(pepd_list_retrieved) ^ set(pepd_id)))	
		elif (n_pepd == len(pepd_id)): 
			if (set(pepd_list_retrieved) != set(pepd_id)):
				print(pepd_list_retrieved)
				print(pepd_id)
				exit('\n\n ... Exiting smoothly ... Missmatch between variants sets (PEPD) \n\n')
			else:
				print(' .. All PEPD entries were collected! .. ')
		
		if (n_neutral < len(neutral_id)):			
			print('\n We are missing %d neutral variant(s) with id(s): ' % (len(neutral_id) - n_neutral))
			print(' %s ' % list(set(neutral_list_retrieved) ^ set(neutral_id)))	
		elif (n_neutral == len(neutral_id)): 
			if (set(neutral_list_retrieved) != set(neutral_id)):
				print(neutral_list_retrieved)
				print(neutral_id)
				exit('\n\n ... Exiting smoothly ... Missmatch between variants sets (NEUTRAL) \n\n')
			else:
				print(' .. All Neutral entries were collected! .. ')			
		
		if (n_lof < len(lof_id)):			
			print('\n We are missing %d lof variant(s) with id(s): ' % (len(lof_id) - n_lof))
			print(' %s ' % list(set(lof_list_retrieved) ^ set(lof_id)))	
		elif (n_lof == len(lof_id)): 
			if (set(lof_list_retrieved) != set(lof_id)):
				print(lof_list_retrieved)
				print(lof_id)
				exit('\n\n ... Exiting smoothly ... Missmatch between variants sets (LoF) \n\n')
			else:
				print('\n\n .. All LoF entries were collected! .. \n\n')	

		if (n_miss < len(miss_id)):			
			print('\n We are missing %d miss variant(s) with id(s): ' % (len(miss_id) - n_miss))
			print(' %s ' % list(set(miss_list_retrieved) ^ set(miss_id)))	
		elif (n_miss == len(miss_id)): 
			if (set(miss_list_retrieved) != set(miss_id)):
				print(miss_list_retrieved)
				print(miss_id)
				exit('\n\n ... Exiting smoothly ... Missmatch between variants sets (missclassified) \n\n')
			else:
				print('\n\n .. All miss entries were collected! .. \n\n')
		
		if (n_gnomad < len(gnomad_id)):			
			print('\n We are missing %d (out of %d) gnomad variant(s). ' % ((len(gnomad_id) - n_gnomad), len(gnomad_id)) )
			# print(' %s ' % list(set(gnomad_list_retrieved) ^ set(gnomad_id)))	
		elif (n_gnomad == len(gnomad_id)): 
			if (set(gnomad_list_retrieved) != set(gnomad_id)):
				print(gnomad_list_retrieved)
				print(gnomad_id)
				exit('\n\n ... Exiting smoothly ... Missmatch between variants sets (gnomad) \n\n')
			else:
				print('\n\n .. All Gnomad entries were collected! .. \n\n')
		
		
		# Benign
		print(' We got %d benign from gnomad ' % n_benign_gnomad)
		print(' We got %d benign/likely benign from gnomad ' % n_benignLikelyBenign_gnomad)
		print(' We got %d likely benign from gnomad ' % n_likelyBenign_gnomad)
		# Pathogenic
		print(' We got %d pathogenic from gnomad ' % n_pathogenic_gnomad)
		print(' We got %d likely benign/likely benign from gnomad ' % n_pathogenicLikelyPathogenic_gnomad)
		print(' We got %d likely pathogenic from gnomad ' % n_likelyPathogenic_gnomad)
		print(' We got %d pathogenic of confl inter from gnomad ' % n_conflInter_gnomad)
		print(' We got %d vus from gnomad ' % n_vus_gnomad)
		
		if (n_clinvar < len(clinvar_id)):			
			print('\n We are missing %d (out of %d) clinvar variant(s). ' % ((len(clinvar_id) - n_clinvar), len(clinvar_id)) )
			# print(' %s ' % list(set(clinvar_list_retrieved) ^ set(clinvar_id)))	
		elif (n_clinvar == len(clinvar_id)): 
			if (set(clinvar_list_retrieved) != set(clinvar_id)):
				print(clinvar_list_retrieved)
				print(clinvar_id)
				exit('\n\n ... Exiting smoothly ... Missmatch between variants sets (clinvar) \n\n')
			else:
				print('\n\n .. All ClinVar entries were collected! .. \n\n')
		
		# Benign
		print(' We got %d benign from clinvar ' % n_benign_clinvar)
		print(' We got %d likely benign/likely benign from clinvar ' % n_benignLikelyBenign_clinvar)
		print(' We got %d likely benign from clinvar ' % n_likelyBenign_clinvar)
		# Pathogenic
		print(' We got %d pathogenic from clinvar ' % n_pathogenic_clinvar)
		print(' We got %d likely benign/likely benign from clinvar ' % n_pathogenicLikelyPathogenic_clinvar)
		print(' We got %d likely pathogenic from clinvar ' % n_likelyPathogenic_clinvar)
		print(' We got %d pathogenic of confl inter from clinvar ' % n_conflInter_clinvar)
		print(' We got %d vus from clinvar ' % n_vus_clinvar)
		
		Tools.StoreFile(domainBoundaries, 'domainBoundaries')
		Tools.StoreFile(varInfo, 'varInfo')
	
	else:
		print('\n\n There is no variants information to be analyzed \n\n')


"""
		Explore the distribution of hotspot locations inside your channel 
"""
def MutationDistribution(mutationSubset, percenctInt = 25, PLOT_ENTROPY_LINE = False):

	## Get orientation	
	mol_info_fn = glob.glob('*_molInfo.xlsx')[0]
	wb = openpyxl.load_workbook(mol_info_fn)	 
	source = wb['geometry_information']
	ind_ppAxis	= source.cell(row = 2, column = 1).value 
	axis_ori = np.sign(ind_ppAxis) 
	ind_ppAxis = abs(ind_ppAxis) - 1
	scales = Tools.LoadFile('scales')
	nrOfAtoms = Tools.LoadFile('nrOfAtoms')

	## Load molecular information
	mol = Tools.LoadFile('mol')

	## Get principal pore axis coord
	princ_coord	= mol['porePoints'][:,ind_ppAxis]*axis_ori
	nrOfPps = len(princ_coord)

	## Load variants	
	varInfo = Tools.LoadFile('varInfo')
	varInfo_res = Tools.GetColumn(varInfo, [2])
	varInfo_type = Tools.GetColumn(varInfo, [1])
	
	from collections import Counter

	# List of amino acids to check
	res = ['ARG', 'MET', 'ALA', 'CYS', 'ASP', 'GLU', 'PHE', 'GLY', 'HIS', 'ILE', 'LEU', 'LYS', 'ASN', 'PRO', 'GLN', 'SER', 'THR', 'TRP', 'TYR', 'VAL']
	# Count how many times each amino acid appears
	count = Counter(varInfo_res)
	# Check how many times each amino acid from the first list appears in the second list
	amino_acid_counts = {aa: count.get(aa, 0) for aa in res}
	
	# Hydropathic scores for each residue (Kapcha & Rossky scale)
	hydropathic_scores = {
    "PHE": -4.0, "PRO": -3.0, "ILE": -3.5, "LEU": -3.5, "TRP": -3.0,
    "VAL": -2.0, "TYR": -1.5, "MET": -1.0, "ALA": 1.0, "THR": 2.0,
    "GLY": 2.5, "CYS": 3.5, "SER": 3.5, "GLN": 5.0, "HIS": 6.0,
    "LYS": 5.0, "GLU": 6.0, "ASN": 6.5, "ASP": 7.5, "ARG": 14.5
	}

	# Normalize counts
	total_count = sum(amino_acid_counts.values())
	normalized_counts = {
    residue: count / total_count for residue, count in amino_acid_counts.items()
	}

	# Normalize hydropathic scores for coloring
	min_score = min(hydropathic_scores.values())
	max_score = max(hydropathic_scores.values())
	normalized_scores = {
    residue: (score - min_score) / (max_score - min_score)
    for residue, score in hydropathic_scores.items()
	}

	# Sort residues from most hydrophobic to most hydrophilic based on hydropathic scores
	sorted_residues = sorted(hydropathic_scores.keys(), key=lambda x: hydropathic_scores[x])

	# Prepare sorted data for plotting
	sorted_counts = [normalized_counts[residue] for residue in sorted_residues]
	sorted_colors = [
    plt.cm.coolwarm(normalized_scores[residue])
    for residue in sorted_residues
	]
	sorted_formatted_residues = [residue.capitalize() for residue in sorted_residues]

	# Plot the bar chart with sorted residues
	plt.figure(figsize=(12, 8))
	bars = plt.bar(sorted_formatted_residues, sorted_counts, color=sorted_colors)
	plt.xlabel('Residue', fontsize = 50)
	plt.ylabel('Freq.', fontsize = 50)
	# plt.title('Kapcha$\&$ Rossky hydropathic scale')
	plt.xticks(rotation=90)

	# Add a color bar to indicate hydrophobic (blue) to hydrophilic (red)
	sm = plt.cm.ScalarMappable(cmap='coolwarm', norm=plt.Normalize(vmin=min_score, vmax=max_score))
	sm.set_array([])
	cbar = plt.colorbar(sm)
	cbar.set_label('Kapcha$\&$Rossky hydropathic scale', fontsize=35)  # Adjust fontsize here

	# Adjust font size of colorbar labels
	cbar.ax.tick_params(labelsize = 20)
	# cbar.ax.tick_params(labelsize=12)

	# Display the plot
	plt.tight_layout()
	plt.show()

	## Get geom model parameters
	statMod = Tools.LoadFile('statMod')
	A = Tools.GetColumn(statMod, [0, 0])	
	a = Tools.GetColumn(statMod, [0, 2])
	l_i = Tools.GetColumn(statMod, [0, 4])
	nu 	= Tools.GetColumn(statMod, [0, 6])	
	modType = Tools.GetColumn(statMod, [0, 15])

	## Get variants indices
	inds_mutations = Tools.GroupVariants(varInfo_type, mutationSubset[0])
	inds_mutations_compare = Tools.GroupVariants(varInfo_type, mutationSubset[1])
	print("\n The mutation number ratio is: ", len(inds_mutations)/len(inds_mutations_compare))
	
	## Initialize
	stats_dist_l_i = []
	dists_l_i = []
	dists_l_i_b = []
	dists_l_half_model = []
	dists_l_half_empirical = []
	l_half_model = np.zeros(nrOfPps)
	l_half_empirical = np.zeros(nrOfPps)
	median_l_mut = np.zeros(nrOfPps)
	mode_l_mut = np.zeros(nrOfPps)
	l_i_b = np.zeros(nrOfPps)

	for i in range(nrOfPps):

		l = scales[i]

		pp = mol['porePoints'][i,:]
		l_res = Tools.EuclNorm(mol['coords_res'], pp)
		
		# Calculate coin-tossing probability
		n_atoms = nrOfAtoms[i][0]
		n_model = Tools.GeomModel(l, [A[i],a[i],l_i[i],nu[i]], modType[i]) * max(n_atoms)
		# Define the coin tossing probability
		# Model:
		p_coin_model = n_model / (max(n_model)) 
		l_i_b[i] = n_model / (A[i] * max(n_atoms))
		p_coin_empirical = n_atoms / (max(n_atoms)) 
		absDiff_model_half_model = abs(p_coin_model - 0.5)
		absDiff_model_half_empirical = abs(p_coin_empirical - 0.5)
		l_half_model[i] = l[np.where(absDiff_model_half_model == min(absDiff_model_half_model))[0][0]] 
		l_half_empirical[i] = l[np.where(absDiff_model_half_empirical == min(absDiff_model_half_empirical))[0][0]] 
		dist_l_i = l_res - l_i[i] 
		dist_l_i_b = l_res - l_i_b[i] 
		dist_l_half_model = l_res - l_half_model[i] 
		dist_l_half_empirical = l_res - l_half_empirical[i] 
		median_l_mut[i] = np.median(l_res[inds_mutations])
		mode_l_mut[i] = Tools.Mode(l_res[inds_mutations])
		stats_dist_l_i.append(Tools.StatsCalc(dist_l_i[inds_mutations], 0, percenctInt))
		dists_l_i.append(dist_l_i[inds_mutations])
		dists_l_i_b.append(dist_l_i_b[inds_mutations])
		dists_l_half_model.append(dist_l_half_model[inds_mutations])
		dists_l_half_empirical.append(dist_l_half_empirical[inds_mutations])

	# plt.plot(l_half_empirical, "g")
	# plt.plot(l_i, "m")
	plt.plot(l_i_b, "m--")
	# plt.plot(median_l_mut, "k")
	# plt.plot(mode_l_mut, "k--")
	plt.show()

	import seaborn as sns
	from collections import defaultdict
	import pandas as pd

	# Create a dictionary to store distances for each residue
	residue_distances = defaultdict(list)
	residues = varInfo_res
	dists = np.abs(dists_l_half_empirical)

	# Iterate over each row of distances
	for dist_row in dists:
		for residue, dist in zip(residues, dist_row):
			residue_distances[residue].append(dist)

	# Initialize lists to store the median and percentiles
	medians = {}
	absMedians = {}
	percentiles_5_95 = {}

	# Calculate the median and the 5-95 percentiles for each residue
	for residue, distances in residue_distances.items():
		median = np.median(distances)
		percentile_5 = np.percentile(distances, 25)
		percentile_95 = np.percentile(distances, 75)
		medians[residue] = median
		absMedians[residue] = np.abs(median)
		percentiles_5_95[residue] = (percentile_5, percentile_95)

	# Sort residues based on ascending absMedian of distances
	sorted_residues = sorted(absMedians, key=lambda x: absMedians[x])

	# Create a color map based on hydropathy
	cmap = plt.get_cmap('coolwarm')
	norm = plt.Normalize(vmin=min(hydropathic_scores.values()), vmax=max(hydropathic_scores.values()))

	# Prepare the data for plotting in a DataFrame
	import itertools
	plot_data = pd.DataFrame({
		'Residue': list(itertools.chain(*[[residue.capitalize()] * len(residue_distances[residue]) for residue in sorted_residues])),
		'Distance': [dist for residue in sorted_residues for dist in residue_distances[residue]],
		'Hydropathy': [hydropathic_scores[residue] for residue in sorted_residues for _ in residue_distances[residue]]
	})

	# Create the box plot
	plt.figure(figsize=(10, 6))

	# Create the box plot for each residue, ordered by ascending median of distances
	sns.boxplot(
		data=plot_data,
		x='Residue',  # Residue names on the x-axis
		y='Distance',  # Distances on the y-axis
		hue='Hydropathy',  # Use hydropathy scores for coloring
		palette='coolwarm',  # Color gradient
		showmeans=True,  # Show median
		legend=False  # Remove legend
	)

	# Set plot labels and title
	plt.title('Box Plot of Distances by Residue, Ordered by Ascending Median Distance')
	plt.xlabel('Residue')
	plt.ylabel('Distance')
	plt.xticks(rotation=90)

	# Add a color bar for hydropathy scale
	sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
	sm.set_array([])
	plt.colorbar(sm, label='Hydropathy')
	plt.tight_layout()
	plt.show()
	 
	Tools.Plot_DistanceFromInflectionPoints(princ_coord, stats_dist_l_i, dists_l_i, l_half_model - l_i, l_half_empirical - l_i, mol['poreRadius_mean'], mol['poreRadius_std'], title = mutationSubset[0], PLOT_ENTROPY_LINE = PLOT_ENTROPY_LINE)
	

'''
		Extract feature: calcualte mutated \phi value
'''
def FeaturesExtraction(classes, classes_unseen, missclassified, jth_order, percenctInt = 25, subClasses = [], VUS = False, PLOT = False, statMethod = "MEDIAN", perp_coord_plot = []):

	## Get orientation	
	mol_info_fn = glob.glob('*_molInfo.xlsx')[0]
	wb = openpyxl.load_workbook(mol_info_fn)	 
	source = wb['geometry_information']
	ind_ppAxis	= source.cell(row = 2, column = 1).value 
	axis_ori = np.sign(ind_ppAxis) 
	ind_ppAxis = abs(ind_ppAxis) - 1
	scales = Tools.LoadFile('scales')

	## Load molecular information
	mol = Tools.LoadFile('mol')

	## Get principal pore axis coord
	princ_coord	= mol['porePoints'][:,ind_ppAxis]*axis_ori
	nrOfPps = len(princ_coord)

	## Geom model parameters
	statMod = Tools.LoadFile('statMod')
	l_i = Tools.GetColumn(statMod, [0, 4])	
	A = Tools.GetColumn(statMod, [0, 0])
	a = Tools.GetColumn(statMod, [0, 2])
	l_i = Tools.GetColumn(statMod, [0, 4])
	modType = Tools.GetColumn(statMod, [0, 15])
	# calculate \xi (we need to set the size of the sliding window)
	nu = Tools.GetColumn(statMod, [0, 6])
	l_lag = Tools.GetColumn(statMod, [0, 8])
	 
	## Load hydr moments
	hydrMoments = Tools.LoadFile('hydrMoments')

	## Load variants	
	varInfo = Tools.LoadFile('varInfo')
	varInfo_type = Tools.GetColumn(varInfo, [1])
	varInfo_classificationStatus = Tools.GetColumn(varInfo, [3])

	## Get variants indices
	inds_class0 = Tools.GroupVariants(varInfo_type, classes[0])
	inds_class1 = Tools.GroupVariants(varInfo_type, classes[1])
	inds_unseen0 = Tools.GroupVariants(varInfo_type, classes_unseen[0])
	inds_unseen1 = Tools.GroupVariants(varInfo_type, classes_unseen[1])
	inds_missclassified = Tools.GroupVariants(varInfo_classificationStatus, missclassified[0])
	# append all the indices
	inds_mut = []
	for lst in (inds_class0, inds_class1, inds_unseen0, inds_unseen1):
		for item in lst:
			inds_mut.append(item)
	# Consider also subclasses, and find where their indices appear in inds_mut
	inds_subClasses_in_mut = []
	if (len(subClasses) != 0):
		for i in range(len(subClasses)):
				inds_subClass = Tools.GroupVariants(varInfo_type, subClasses[i])
				inds_subClasses_in_mut.append([j for j, item in enumerate(inds_mut) if item in inds_subClass])
	
	# now append also the missclassified ones at the inds_mut list (to avoid returining double subclass indices!)
	inds_mut = np.append(inds_mut, inds_missclassified)
	
	# varInfo_resID = Tools.GetColumn(varInfo, [0])
	# print(varInfo_resID[inds_missclassified])
	# Include VUS
	if (VUS == True):
		inds_vus = Tools.GroupVariants(varInfo_type, "VUS")
		n_class_vus = len(inds_vus)
	# Count ..
	n_class_0 = len(inds_class0)
	n_class_1 = len(inds_class1)
	ratio = n_class_0/n_class_1

	print('\n The ratio of %s class instances to %s class instances is %.2f ' % (classes[0], classes[1], ratio))
	print('\n Number of cases considered: %s' % (n_class_0 + n_class_1))

	## Create for accessing the observable arrays
	inds_class0_ = np.arange(0, len(inds_class0), 1)
	inds_class1_ = np.arange(len(inds_class0), len(inds_class0) + len(inds_class1), 1)

	## Initialize
	# Statistical summaries ..
	# everything comes in pairs (except from missclassified ones):
	# class 0
	stats_phi_class0 = []
	stats_derPhi_class0 = []
	stats_pertPot_class0 = []
	stats_absPertPot_class0 = []
	# class 1
	stats_phi_class1 = []
	stats_derPhi_class1 = []
	stats_pertPot_class1 = []
	stats_absPertPot_class1 = []

	## All mutations
	feature_phi = []
	feature_absPhi = []
	feature_derPhi = []
	feature_absDerPhi = []
	feature_pertPot = []
	feature_absPertPot = []
	
	if (len(classes_unseen) != 0):
		# 0
		stats_phi_unseen0 = []
		stats_derPhi_unseen0 = []
		stats_pertPot_unseen0 = []
		# 1
		stats_phi_unseen1 = []
		stats_derPhi_unseen1 = []
		stats_pertPot_unseen1 = []
	if (len(missclassified) != 0):
		stats_phi_missclassified = [] # these are misclasified events
		stats_derPhi_missclassified = [] # these are misclasified events
		stats_pertPot_missclassified = [] # these are misclasified events
	if (len(subClasses) != 0):
		# A
		stats_phi_subClassA = [] 
		stats_derPhi_subClassA = [] 
		stats_pertPot_subClassA = []
		# B 
		stats_phi_subClassB = [] 
		stats_derPhi_subClassB = [] 
		stats_pertPot_subClassB = []

	for i in range(nrOfPps):

		l = scales[i]
		# ind_l_i = np.argmin(abs(l - l_i[i]))
		# ind_l_lag = np.argmin(abs(l - l_lag[i]))
		curvature = Tools.GeomModel(l, [A[i],a[i],l_i[i],nu[i]], modType[i], der=2) / A[i]
		ind_max_curv = Tools.Match(max(curvature), curvature)[0]
		pp = mol['porePoints'][i,:]
		l_res = Tools.EuclNorm(mol['coords_res'], pp)
	
		if (jth_order % 2 == 0):
			h_pho = Tools.getHydrMom(hydrMoments, i, 'phobic', jth_order) + ModelParameters.ZERO # add ZERO to avoid real zeros which are empty sheres.
			h_phi = Tools.getHydrMom(hydrMoments, i, 'philic', jth_order) + ModelParameters.ZERO
		else:
			h_pho = Tools.getHydrMom(hydrMoments, i, 'phobic', jth_order, "z") + ModelParameters.ZERO
			h_phi = Tools.getHydrMom(hydrMoments, i, 'philic', jth_order, "z") + ModelParameters.ZERO

		# Check if h is decomposable 
		# The cutoff scale is chosen to be the l-value for which the n-curavture is maximized
		# This is quite relaxed criterion: a more strict criterion also works. 
		# In fact, we can take l_cutoff to be *very* small.
		h_plus, h_minus = Tools.Decompose(h_pho, h_phi, ind_max_curv)

		# Define phi for increasing l (log of comp susc)
		phi = np.log(abs(h_plus / h_minus))

		# Intrapolate to get the \phi value at l_mut (we may do that since \Delta l is very small)
		l_mut = l_res[inds_mut]
		phi_mut = Tools.Intrapolator(l, phi, l_mut)

		# Consider plotting something here
		if (len(perp_coord_plot) == 2):
			if (pp[2] > perp_coord_plot[0] and pp[2] < perp_coord_plot[1]):
				
				print(pp[2])
				fig, ax = plt.subplots()
				ax.set_title("$\mathrm{p}_{\perp}= 3.1$")
				ax.plot(l_mut[inds_class0_], phi_mut[inds_class0_], "ro", label = "GoF$\cup$LoF")
				ax.plot(l_mut[inds_class1_], phi_mut[inds_class1_], "bo", label = "Neutr.$\cup$Benign")
				ax.plot(l, phi,"k", alpha = 0.5)
				ax.axhline(y=0, color="k", alpha = 0.25, linestyle = "--", linewidth = 5)
				ax.axvline(x = l_i[i], color = "k", linestyle = "--", linewidth = 5, alpha = 0.25)
				ax.legend(loc="upper left", fontsize = 40)
				ax.set_ylabel("$\phi_{1,\perp}$", fontsize = 55)
				ax.set_xlabel("$l$ [\AA]", fontsize = 55)
				ax.set_xticks([10, 20, 30, l_i[i], 50, 60, 70])
				ax.set_xticklabels(["10", "20", "30", "$l_i$", "50", "60", "70"],  fontsize=40)
				inset_ax = fig.add_axes([0.66, 0.58, 0.23, 0.25]) 
				inset_ax.plot(l, h_plus/max(abs(h_plus)), label = "$h_{1,\perp,+}/\mathrm{max}\{|h_{1,\perp,+}|\}$", linewidth = 5)
				inset_ax.plot(l, h_minus/max(abs(h_minus)), label = "$h_{1,\perp,-}/\mathrm{max}\{|h_{1,\perp,-}|\}$", linewidth = 5)
				inset_ax.legend(loc="lower left", fontsize = 17)
				inset_ax.set_ylabel("$h_{1,\perp,\pm}$ [kcal$\cdot$\AA]", fontsize = 30)
				inset_ax.set_xlabel("$l$ [\AA]", fontsize = 30)
				inset_ax.axvline(x = l_i[i], color = "k", linestyle = "--", linewidth = 5, alpha = 0.25)
				inset_ax.axhline(y=0, color="k", alpha = 0.25, linestyle = "--", linewidth = 5)
				inset_ax.set_xticks([10, 20, 30, l_i[i], 50, 60, 70])
				inset_ax.set_xticklabels(["10", "20", "30", "$l_i$", "50", "60", "70"],  fontsize=30)
				inset_ax.set_yticks([-1, 0, 1])
				inset_ax.set_yticklabels(["-1", "0", "1"],  fontsize=30)

				plt.show()

		# Differtiate phi_mut - note that this essential step! How to differeniate while filtering out noise? 
		# Alternatively, one could calculate the scaling exponents of h_plus, h_minus. 
		# However, this approach does not yield high enough accuracy, since exponents are extracted from log-log diagrams,
		# and, hence, oscillatory behavior of the derivative is lost. These oscillations convey important information 
		# about the underlying wave packet.
		# Higher order derivatiges of phi_mut are welcome. However, it is increasingly difficult to 
		# compute a smooth, yet, informative derivative with increasing order.
		slidingWindow = int( ModelParameters.WINDOW_FAC * Tools.GetSlidingWindow(a[i], nu[i], modType[i]) )
		derPhi_mut = Tools.RadialProfile(l, phi, slidingWindow, l_mut, 1)
		pertPot_mut = derPhi_mut / phi_mut
		
		# class 0
		stats_phi_class0.append(Tools.StatsCalc(phi_mut[inds_class0_], 0, percenctInt))
		stats_derPhi_class0.append(Tools.StatsCalc(derPhi_mut[inds_class0_], 0, percenctInt))
		stats_pertPot_class0.append(Tools.StatsCalc(pertPot_mut[inds_class0_], 0, percenctInt))
		stats_absPertPot_class0.append(Tools.StatsCalc(abs(pertPot_mut[inds_class0_]), 0, percenctInt))
		# class 1
		stats_phi_class1.append(Tools.StatsCalc(phi_mut[inds_class1_], 0, percenctInt))
		stats_derPhi_class1.append(Tools.StatsCalc(derPhi_mut[inds_class1_], 0, percenctInt))
		stats_pertPot_class1.append(Tools.StatsCalc(pertPot_mut[inds_class1_], 0, percenctInt))
		stats_absPertPot_class1.append(Tools.StatsCalc(abs(pertPot_mut[inds_class1_]), 0, percenctInt))

		if (len(classes_unseen) != 0):
			inds_unseen0_ = np.arange(len(inds_class0) + len(inds_class1), len(inds_class0) + len(inds_class1) + len(inds_unseen0), 1)
			inds_unseen1_ = np.arange(len(inds_class0) + len(inds_class1) + len(inds_unseen0), len(inds_class0) + len(inds_class1) + len(inds_unseen0) + len(inds_unseen1), 1)
			
			# class unseen 0
			stats_phi_unseen0.append(Tools.StatsCalc(phi_mut[inds_unseen0_], 0, percenctInt))
			stats_derPhi_unseen0.append(Tools.StatsCalc(derPhi_mut[inds_unseen0_], 0, percenctInt))
			stats_pertPot_unseen0.append(Tools.StatsCalc(pertPot_mut[inds_unseen0_], 0, percenctInt))
			# class unseen 1
			stats_phi_unseen1.append(Tools.StatsCalc(phi_mut[inds_unseen1_], 0, percenctInt))
			stats_derPhi_unseen1.append(Tools.StatsCalc(derPhi_mut[inds_unseen1_], 0, percenctInt))
			stats_pertPot_unseen1.append(Tools.StatsCalc(pertPot_mut[inds_unseen1_], 0, percenctInt))
			
		if (len(missclassified) != 0):
			inds_missclassified_ = np.arange(len(inds_class0) + len(inds_class1) + len(inds_unseen0) + len(inds_unseen1), len(inds_class0) + len(inds_class1) + len(inds_unseen0) + len(inds_unseen1) + len(inds_missclassified), 1)

			# missclassified subset
			stats_phi_missclassified.append(Tools.StatsCalc(phi_mut[inds_missclassified_], 0, percenctInt))
			stats_derPhi_missclassified.append(Tools.StatsCalc(derPhi_mut[inds_missclassified_], 0, percenctInt))
			stats_pertPot_missclassified.append(Tools.StatsCalc(pertPot_mut[inds_missclassified_], 0, percenctInt))
		
		## stats for subclasses
		if (len(subClasses) != 0):
			# A subset
			inds_A = inds_subClasses_in_mut[0]
			stats_phi_subClassA.append(Tools.StatsCalc(phi_mut[inds_A], 0, percenctInt)) 
			stats_derPhi_subClassA.append(Tools.StatsCalc(derPhi_mut[inds_A], 0, percenctInt)) 
			stats_pertPot_subClassA.append(Tools.StatsCalc(pertPot_mut[inds_A], 0, percenctInt)) 
			# B subset
			inds_B = inds_subClasses_in_mut[1]
			stats_phi_subClassB.append(Tools.StatsCalc(phi_mut[inds_B], 0, percenctInt)) 
			stats_derPhi_subClassB.append(Tools.StatsCalc(derPhi_mut[inds_B], 0, percenctInt)) 
			stats_pertPot_subClassB.append(Tools.StatsCalc(pertPot_mut[inds_B], 0, percenctInt)) 

		## Append candidate features!
		feature_phi.append(phi_mut)
		feature_absPhi.append(np.abs(phi_mut))
		feature_derPhi.append(derPhi_mut)
		feature_absDerPhi.append(np.abs(derPhi_mut))
		feature_pertPot.append(pertPot_mut)
		feature_absPertPot.append(abs(pertPot_mut))
		
	"""
	plt.hist(np.array(feature_phi).flatten().tolist(), bins = 50, edgecolor = "b", alpha = 0.5, label = "$\phi_{1,\perp}$")
	plt.hist(np.array(feature_derPhi).flatten().tolist(), bins = 50, edgecolor = "r", alpha = 0.5, label = "$\mathcal{I}_{1,\perp}$")
	plt.legend(loc = "upper right", fontsize = 50)
	plt.xlabel("$\phi_{\perp}$, $\mathcal{I}_{\perp}$", fontsize = 50)
	plt.ylabel("Freq.",fontsize = 50)
	plt.show()
	"""

	if (PLOT == True):
		# phi	
		index_name = r"\( \phi_{"+ str(jth_order) +"} \)"
		Tools.Plot_FeatureStatistic(princ_coord, stats_phi_class0, stats_phi_class1, mol['poreRadius_mean'], mol['poreRadius_std'], classes[0], classes[1], statMethod, index_name)
		# derPhi
		index_name = r"\( \mathcal{I}_{"+ str(jth_order) +"} \)"
		Tools.Plot_FeatureStatistic(princ_coord, stats_derPhi_class0, stats_derPhi_class1, mol['poreRadius_mean'], mol['poreRadius_std'], classes[0], classes[1], statMethod, index_name)
		# pertPot
		# index_name = r"\( \mathcal{I}_{"+ str(jth_order) +"} / \phi_{"+ str(jth_order) +"}  \)"
		# Tools.Plot_FeatureStatistic(princ_coord, stats_pertPot_class0, stats_pertPot_class1, mol['poreRadius_mean'], mol['poreRadius_std'], classes[0], classes[1], statMethod, index_name)
		# absPertPot
		# index_name = r"\( | \mathcal{I}_{"+ str(jth_order) +"} / \phi_{"+ str(jth_order) +"} | \)"
		# Tools.Plot_FeatureStatistic(princ_coord, stats_absPertPot_class0, stats_absPertPot_class1, mol['poreRadius_mean'], mol['poreRadius_std'], classes[0], classes[1], statMethod, index_name)
		# phi subClass_0 vs subClass_1
		if (len(subClasses) != 0):
	
			index_name = r"\( \phi_{"+ str(jth_order) +"} \)"
			Tools.Plot_FeatureStatistic(princ_coord, stats_phi_subClassA, stats_phi_subClassB, mol['poreRadius_mean'], mol['poreRadius_std'], subClasses[0], subClasses[1], statMethod, index_name)

			index_name = r"\( \mathcal{I}_{"+ str(jth_order) +"} \)"
			Tools.Plot_FeatureStatistic(princ_coord, stats_derPhi_subClassA, stats_derPhi_subClassB, mol['poreRadius_mean'], mol['poreRadius_std'], subClasses[0], subClasses[1], statMethod, index_name)
	
	# Save candidate features
	# retrieve median of each residue with: np.median(feature_, axis = 0)
	fn = str(jth_order) + '-order_phi'
	Tools.StoreFile(feature_phi, fn)
	fn = str(jth_order) + '-order_absPhi'
	Tools.StoreFile(feature_absPhi, fn)
	fn = str(jth_order) + '-order_derPhi'
	Tools.StoreFile(feature_derPhi, fn)
	fn = str(jth_order) + '-order_absDerPhi'
	Tools.StoreFile(feature_absDerPhi, fn)
	fn = str(jth_order) + '-order_pertPot'
	Tools.StoreFile(feature_pertPot, fn)
	fn = str(jth_order) + '-order_absPertPot'
	Tools.StoreFile(feature_absPertPot, fn)

	# Store only the medians the left and right percentiles of candidate features
	# Store classes 0,1 
	fn = str(jth_order) + '-order_medians_classes'
	Tools.StoreFile(	
						[	
							# 0
							Tools.GetColumn(stats_phi_class0, [2]), Tools.GetColumn(stats_phi_class0, [3]), Tools.GetColumn(stats_phi_class0, [4]), 		# 0,1,2
							Tools.GetColumn(stats_derPhi_class0, [2]), Tools.GetColumn(stats_derPhi_class0, [3]), Tools.GetColumn(stats_derPhi_class0, [4]), 	# 3,4,5	
							Tools.GetColumn(stats_pertPot_class0, [2]), Tools.GetColumn(stats_pertPot_class0, [3]), Tools.GetColumn(stats_pertPot_class0, [4]), 	# 6,7,8	
							Tools.GetColumn(stats_absPertPot_class0, [2]), Tools.GetColumn(stats_absPertPot_class0, [3]), Tools.GetColumn(stats_absPertPot_class0, [4]), # 9,10,11 		
				  			# 1
							Tools.GetColumn(stats_phi_class1, [2]), Tools.GetColumn(stats_phi_class1, [3]), Tools.GetColumn(stats_phi_class1, [4]), 		# 12,13,14
							Tools.GetColumn(stats_derPhi_class1, [2]), Tools.GetColumn(stats_derPhi_class1, [3]), Tools.GetColumn(stats_derPhi_class1, [4]), 	# 15,16,17	
							Tools.GetColumn(stats_pertPot_class1, [2]), Tools.GetColumn(stats_pertPot_class1, [3]), Tools.GetColumn(stats_pertPot_class1, [4]), 	# 18,19,20	
							Tools.GetColumn(stats_absPertPot_class1, [2]), Tools.GetColumn(stats_absPertPot_class1, [3]), Tools.GetColumn(stats_absPertPot_class1, [4]), # 21,22,23 
						], 
					fn)
	
	# Store the unseens
	if (len(classes_unseen) != 0):
		fn = str(jth_order) + '-order_medians_unseen'
		Tools.StoreFile(	
						[	
							# 0
							Tools.GetColumn(stats_phi_unseen0, [2]), Tools.GetColumn(stats_phi_unseen0, [3]), Tools.GetColumn(stats_phi_unseen0, [4]), 		
							Tools.GetColumn(stats_derPhi_unseen0, [2]), Tools.GetColumn(stats_derPhi_unseen0, [3]), Tools.GetColumn(stats_derPhi_unseen0, [4]), 		
							Tools.GetColumn(stats_pertPot_unseen0, [2]), Tools.GetColumn(stats_pertPot_unseen0, [3]), Tools.GetColumn(stats_pertPot_unseen0, [4]), 				
				  			# 1
							Tools.GetColumn(stats_phi_unseen1, [2]), Tools.GetColumn(stats_phi_unseen1, [3]), Tools.GetColumn(stats_phi_unseen1, [4]), 		
							Tools.GetColumn(stats_derPhi_unseen1, [2]), Tools.GetColumn(stats_derPhi_unseen1, [3]), Tools.GetColumn(stats_derPhi_unseen1, [4]), 		
							Tools.GetColumn(stats_pertPot_unseen1, [2]), Tools.GetColumn(stats_pertPot_unseen1, [3]), Tools.GetColumn(stats_pertPot_unseen1, [4]), 		
						], 
					fn)

	# Store subclasses
	if (len(subClasses) != 0):
		fn = str(jth_order) + '-order_medians_subclasses'
		Tools.StoreFile(	
						[	
							# A
							Tools.GetColumn(stats_phi_subClassA, [2]), Tools.GetColumn(stats_phi_subClassA, [3]), Tools.GetColumn(stats_phi_subClassA, [4]), 		
							Tools.GetColumn(stats_derPhi_subClassA, [2]), Tools.GetColumn(stats_derPhi_subClassA, [3]), Tools.GetColumn(stats_derPhi_subClassA, [4]), 		
							Tools.GetColumn(stats_pertPot_subClassA, [2]), Tools.GetColumn(stats_pertPot_subClassA, [3]), Tools.GetColumn(stats_pertPot_subClassA, [4]), 				
				  			# B
							Tools.GetColumn(stats_phi_subClassB, [2]), Tools.GetColumn(stats_phi_subClassB, [3]), Tools.GetColumn(stats_phi_subClassB, [4]), 		
							Tools.GetColumn(stats_derPhi_subClassB, [2]), Tools.GetColumn(stats_derPhi_subClassB, [3]), Tools.GetColumn(stats_derPhi_subClassB, [4]), 		
							Tools.GetColumn(stats_pertPot_subClassB, [2]), Tools.GetColumn(stats_pertPot_subClassB, [3]), Tools.GetColumn(stats_pertPot_subClassB, [4]), 		
						], 
					fn)
		
	# Store missclassified
	if (len(missclassified) != 0):
		fn = str(jth_order) + '-order_medians_missclassified'
		Tools.StoreFile(	
						[	
							Tools.GetColumn(stats_phi_missclassified, [2]), Tools.GetColumn(stats_phi_missclassified, [3]), Tools.GetColumn(stats_phi_missclassified, [4]), 		
							Tools.GetColumn(stats_derPhi_missclassified, [2]), Tools.GetColumn(stats_derPhi_missclassified, [3]), Tools.GetColumn(stats_derPhi_missclassified, [4]), 		
							Tools.GetColumn(stats_pertPot_missclassified, [2]), Tools.GetColumn(stats_pertPot_missclassified, [3]), Tools.GetColumn(stats_pertPot_missclassified, [4]), 						
						], 
					fn)


"""
		Features summary (in terms of median statistics)
"""
def FeaturesSummary(even_order, odd_order, classes = 'classes', unseen = 'unseen', missclassified = 'missclassified', subclasses = 'subclasses', 
					label_class0 = r"GoF\( \cup \)LoF", label_class1 = "Neut.", label_unseen1 = "Benign", label_missclass = "missclass."):

	# Get orientation	
	mol_info_fn = glob.glob('*_molInfo.xlsx')[0]
	wb = openpyxl.load_workbook(mol_info_fn)	 
	source = wb['geometry_information']
	ind_ppAxis	= source.cell(row = 2, column = 1).value 
	axis_ori = np.sign(ind_ppAxis) 
	ind_ppAxis = abs(ind_ppAxis) - 1

	# Load molecular information
	mol = Tools.LoadFile('mol')

	# Get principal pore axis coord
	princ_coord	= mol['porePoints'][:,ind_ppAxis]*axis_ori
	
	# Even contributions
	# medians 0,1
	medians_even_class0_phi = []
	medians_even_class0_derPhi = []
	medians_even_class0_pertPot = []
	medians_even_class0_absPertPot = []

	medians_even_class1_phi = []
	leftPerc_even_class1_phi = []
	rightPerc_even_class1_phi = []

	medians_even_class1_derPhi = []
	leftPerc_even_class1_derPhi = []
	rightPerc_even_class1_derPhi = []

	medians_even_class1_pertPot = []
	medians_even_class1_absPertPot = []

	for j in even_order:
		fn = str(j) + '-order_medians_' + classes
		# 0
		medians_even_class0_phi.append(Tools.LoadFile(fn)[0])
		medians_even_class0_derPhi.append(Tools.LoadFile(fn)[3])
		medians_even_class0_pertPot.append(Tools.LoadFile(fn)[6])
		medians_even_class0_absPertPot.append(Tools.LoadFile(fn)[9])
		# 1
		medians_even_class1_phi.append(Tools.LoadFile(fn)[12])
		leftPerc_even_class1_phi.append(Tools.LoadFile(fn)[13])
		rightPerc_even_class1_phi.append(Tools.LoadFile(fn)[14])

		medians_even_class1_derPhi.append(Tools.LoadFile(fn)[15])
		leftPerc_even_class1_derPhi.append(Tools.LoadFile(fn)[16])
		rightPerc_even_class1_derPhi.append(Tools.LoadFile(fn)[17])
		medians_even_class1_pertPot.append(Tools.LoadFile(fn)[18])
		medians_even_class1_absPertPot.append(Tools.LoadFile(fn)[21])

	# Odd contributions
	# medians 0,1
	medians_odd_class0_phi = []
	medians_odd_class0_derPhi = []

	medians_odd_class0_pertPot = []
	medians_odd_class0_absPertPot = []
	
	medians_odd_class1_phi = []
	leftPerc_odd_class1_phi = []
	rightPerc_odd_class1_phi = []

	medians_odd_class1_derPhi = []
	leftPerc_odd_class1_derPhi = []
	rightPerc_odd_class1_derPhi = []

	medians_odd_class1_pertPot = []
	medians_odd_class1_absPertPot = []

	for j in odd_order:
		fn = str(j) + '-order_medians_' + classes
		# 0
		medians_odd_class0_phi.append(Tools.LoadFile(fn)[0])
		
		medians_odd_class0_derPhi.append(Tools.LoadFile(fn)[3])
		medians_odd_class0_pertPot.append(Tools.LoadFile(fn)[6])
		medians_odd_class0_absPertPot.append(Tools.LoadFile(fn)[9])
		# 1
		medians_odd_class1_phi.append(Tools.LoadFile(fn)[12])
		leftPerc_odd_class1_phi.append(Tools.LoadFile(fn)[13])
		rightPerc_odd_class1_phi.append(Tools.LoadFile(fn)[14])
		
		medians_odd_class1_derPhi.append(Tools.LoadFile(fn)[15])
		leftPerc_odd_class1_derPhi.append(Tools.LoadFile(fn)[15])
		rightPerc_odd_class1_derPhi.append(Tools.LoadFile(fn)[15])
		
		medians_odd_class1_pertPot.append(Tools.LoadFile(fn)[18])
		medians_odd_class1_absPertPot.append(Tools.LoadFile(fn)[21])

	
	## Are there any unseen data?
	if (len(unseen) != 0):

		# Even contributions
		# medians 0,1
		medians_even_unseen0_phi = []
		medians_even_unseen0_derPhi = []
		medians_even_unseen0_pertPot = []
		
		medians_even_unseen1_phi = []
		leftPerc_even_unseen1_phi = []
		rightPerc_even_unseen1_phi = []

		medians_even_unseen1_derPhi = []
		leftPerc_even_unseen1_derPhi = []
		rightPerc_even_unseen1_derPhi = []

		medians_even_unseen1_pertPot = []
		
		for j in even_order:
			fn = str(j) + '-order_medians_' + unseen
			# 0
			medians_even_unseen0_phi.append(Tools.LoadFile(fn)[0])
			medians_even_unseen0_derPhi.append(Tools.LoadFile(fn)[3])
			medians_even_unseen0_pertPot.append(Tools.LoadFile(fn)[6])
			
			# 1
			medians_even_unseen1_phi.append(Tools.LoadFile(fn)[9])
			leftPerc_even_unseen1_phi.append(Tools.LoadFile(fn)[10])
			rightPerc_even_unseen1_phi.append(Tools.LoadFile(fn)[11])

			medians_even_unseen1_derPhi.append(Tools.LoadFile(fn)[12])
			leftPerc_even_unseen1_derPhi.append(Tools.LoadFile(fn)[13])
			rightPerc_even_unseen1_derPhi.append(Tools.LoadFile(fn)[14])

			medians_even_unseen1_pertPot.append(Tools.LoadFile(fn)[15])
			
		# Odd contributions
		# medians 0,1
		medians_odd_unseen0_phi = []
		medians_odd_unseen0_derPhi = []
		medians_odd_unseen0_pertPot = []
		
		medians_odd_unseen1_phi = []
		leftPerc_odd_unseen1_phi = []
		rightPerc_odd_unseen1_phi = []

		medians_odd_unseen1_derPhi = []
		leftPerc_odd_unseen1_derPhi = []
		rightPerc_odd_unseen1_derPhi = []

		medians_odd_unseen1_pertPot = []
		

		for j in odd_order:
			fn = str(j) + '-order_medians_' + unseen
			# 0
			medians_odd_unseen0_phi.append(Tools.LoadFile(fn)[0])
			medians_odd_unseen0_derPhi.append(Tools.LoadFile(fn)[3])
			medians_odd_unseen0_pertPot.append(Tools.LoadFile(fn)[6])
			
			# 1
			medians_odd_unseen1_phi.append(Tools.LoadFile(fn)[9])
			leftPerc_odd_unseen1_phi.append(Tools.LoadFile(fn)[10])
			rightPerc_odd_unseen1_phi.append(Tools.LoadFile(fn)[11])

			medians_odd_unseen1_derPhi.append(Tools.LoadFile(fn)[12])
			leftPerc_odd_unseen1_derPhi.append(Tools.LoadFile(fn)[13])
			rightPerc_odd_unseen1_derPhi.append(Tools.LoadFile(fn)[14])
			medians_odd_unseen1_pertPot.append(Tools.LoadFile(fn)[15])


	## Are there any missclassified data?
	if (len(missclassified) != 0):

		# Even contributions
		# medians 
		medians_even_missclassified_phi = []
		leftPerc_even_missclassified_phi = []
		rightPerc_even_missclassified_phi = []

		medians_even_missclassified_derPhi = []
		leftPerc_even_missclassified_derPhi = []
		rightPerc_even_missclassified_derPhi = []

		medians_even_missclassified_pertPot = []

		for j in even_order:
			fn = str(j) + '-order_medians_' + missclassified
			#
			medians_even_missclassified_phi.append(Tools.LoadFile(fn)[0])
			leftPerc_even_missclassified_phi.append(Tools.LoadFile(fn)[1])
			rightPerc_even_missclassified_phi.append(Tools.LoadFile(fn)[2])

			medians_even_missclassified_derPhi.append(Tools.LoadFile(fn)[3])
			leftPerc_even_missclassified_derPhi.append(Tools.LoadFile(fn)[4])
			rightPerc_even_missclassified_derPhi.append(Tools.LoadFile(fn)[5])
			
			medians_even_missclassified_pertPot.append(Tools.LoadFile(fn)[6])

		# Odd contributions
		# medians 
		medians_odd_missclassified_phi = []
		leftPerc_odd_missclassified_phi = []
		rightPerc_odd_missclassified_phi = []

		medians_odd_missclassified_derPhi = []
		leftPerc_odd_missclassified_derPhi = []
		rightPerc_odd_missclassified_derPhi = []
		
		medians_odd_missclassified_pertPot = []

		for j in odd_order:
			fn = str(j) + '-order_medians_' + missclassified
			# 
			medians_odd_missclassified_phi.append(Tools.LoadFile(fn)[0])
			leftPerc_odd_missclassified_phi.append(Tools.LoadFile(fn)[1])
			rightPerc_odd_missclassified_phi.append(Tools.LoadFile(fn)[2])
			
			medians_odd_missclassified_derPhi.append(Tools.LoadFile(fn)[3])
			leftPerc_odd_missclassified_derPhi.append(Tools.LoadFile(fn)[4])
			rightPerc_odd_missclassified_derPhi.append(Tools.LoadFile(fn)[5])

			medians_odd_missclassified_pertPot.append(Tools.LoadFile(fn)[6])

	## Are there any subclasses data?
	if (len(subclasses) != 0):

		# Even contributions
		# medians 
		medians_even_subclassesA_phi = []
		leftPerc_even_subclassesA_phi = []
		rightPerc_even_subclassesA_phi = []

		medians_even_subclassesA_derPhi = []
		leftPerc_even_subclassesA_derPhi = []
		rightPerc_even_subclassesA_derPhi = []

		medians_even_subclassesA_pertPot = []

		medians_even_subclassesB_phi = []
		leftPerc_even_subclassesB_phi = []
		rightPerc_even_subclassesB_phi = []
		
		medians_even_subclassesB_derPhi = []
		leftPerc_even_subclassesB_derPhi = []
		rightPerc_even_subclassesB_derPhi = []
		
		medians_even_subclassesB_pertPot = []

		for j in even_order:
			fn = str(j) + '-order_medians_' + subclasses
			# A
			medians_even_subclassesA_phi.append(Tools.LoadFile(fn)[0])
			leftPerc_even_subclassesA_phi.append(Tools.LoadFile(fn)[1])
			rightPerc_even_subclassesA_phi.append(Tools.LoadFile(fn)[2])

			medians_even_subclassesA_derPhi.append(Tools.LoadFile(fn)[3])
			leftPerc_even_subclassesA_derPhi.append(Tools.LoadFile(fn)[4])
			rightPerc_even_subclassesA_derPhi.append(Tools.LoadFile(fn)[5])

			medians_even_subclassesA_pertPot.append(Tools.LoadFile(fn)[6])
			# B
			medians_even_subclassesB_phi.append(Tools.LoadFile(fn)[9])
			leftPerc_even_subclassesB_phi.append(Tools.LoadFile(fn)[10])
			rightPerc_even_subclassesB_phi.append(Tools.LoadFile(fn)[11])

			medians_even_subclassesB_derPhi.append(Tools.LoadFile(fn)[12])
			leftPerc_even_subclassesB_derPhi.append(Tools.LoadFile(fn)[13])
			rightPerc_even_subclassesB_derPhi.append(Tools.LoadFile(fn)[14])

			medians_even_subclassesB_pertPot.append(Tools.LoadFile(fn)[15])

		# Odd contributions
		# medians 
		medians_odd_subclassesA_phi = []
		leftPerc_odd_subclassesA_phi = []
		rightPerc_odd_subclassesA_phi = []

		medians_odd_subclassesA_derPhi = []
		leftPerc_odd_subclassesA_derPhi = []
		rightPerc_odd_subclassesA_derPhi = []
		
		medians_odd_subclassesA_pertPot = []

		medians_odd_subclassesB_phi = []
		leftPerc_odd_subclassesB_phi = []
		rightPerc_odd_subclassesB_phi = []
		
		medians_odd_subclassesB_derPhi = []
		leftPerc_odd_subclassesB_derPhi = []
		rightPerc_odd_subclassesB_derPhi = []

		medians_odd_subclassesB_pertPot = []

		for j in odd_order:
			fn = str(j) + '-order_medians_' + subclasses
			# A
			medians_odd_subclassesA_phi.append(Tools.LoadFile(fn)[0])
			leftPerc_odd_subclassesA_phi.append(Tools.LoadFile(fn)[1])
			rightPerc_odd_subclassesA_phi.append(Tools.LoadFile(fn)[2])
			
			medians_odd_subclassesA_derPhi.append(Tools.LoadFile(fn)[3])
			leftPerc_odd_subclassesA_derPhi.append(Tools.LoadFile(fn)[4])
			rightPerc_odd_subclassesA_derPhi.append(Tools.LoadFile(fn)[5])
			medians_odd_subclassesA_pertPot.append(Tools.LoadFile(fn)[6])
			# B
			medians_odd_subclassesB_phi.append(Tools.LoadFile(fn)[9])
			leftPerc_odd_subclassesB_phi.append(Tools.LoadFile(fn)[10])
			rightPerc_odd_subclassesB_phi.append(Tools.LoadFile(fn)[11])

			medians_odd_subclassesB_derPhi.append(Tools.LoadFile(fn)[12])
			leftPerc_odd_subclassesB_derPhi.append(Tools.LoadFile(fn)[13])
			rightPerc_odd_subclassesB_derPhi.append(Tools.LoadFile(fn)[14])
			medians_odd_subclassesB_pertPot.append(Tools.LoadFile(fn)[15])
			
			

	##
	## Phi plot (even contributions) ##
	##
	percentiles = [leftPerc_even_subclassesA_phi, rightPerc_even_subclassesA_phi,
					leftPerc_even_subclassesB_phi, rightPerc_even_subclassesB_phi,
					leftPerc_even_class1_phi, rightPerc_even_class1_phi,
					leftPerc_even_unseen1_phi, rightPerc_even_unseen1_phi,
					leftPerc_even_missclassified_phi, rightPerc_even_missclassified_phi]
				
	Tools.Plot_MediansOfFeatureMedias(princ_coord, [medians_even_subclassesA_phi, medians_even_subclassesB_phi], medians_even_class1_phi, mol['poreRadius_mean'], mol['poreRadius_std'], 
								   	 	medians_unseen = medians_even_unseen1_phi, medians_missclass = medians_even_missclassified_phi, SHOW_LEGENG=True)
	
	##
	## Phi plot (odd contributions) ##
	##
	percentiles = [leftPerc_odd_subclassesA_phi, rightPerc_odd_subclassesA_phi,
					leftPerc_odd_subclassesB_phi, rightPerc_odd_subclassesB_phi,
					leftPerc_odd_class1_phi, rightPerc_odd_class1_phi,
					leftPerc_odd_unseen1_phi, rightPerc_odd_unseen1_phi,
					leftPerc_odd_missclassified_phi, rightPerc_odd_missclassified_phi]
	
	Tools.Plot_MediansOfFeatureMedias(princ_coord, [medians_odd_subclassesA_phi, medians_odd_subclassesB_phi], medians_odd_class1_phi, mol['poreRadius_mean'], mol['poreRadius_std'], 
								   		medians_unseen = medians_odd_unseen1_phi, medians_missclass = medians_odd_missclassified_phi, y_label = 'Interfacial inertia $\mathbf{p}$' )  # y_label = 'Cluster conductivity')
	

	##
	## derPhi plot (even contributions) ##
	##
	percentiles = [leftPerc_even_subclassesA_derPhi, rightPerc_even_subclassesA_derPhi,
					leftPerc_even_subclassesB_derPhi, rightPerc_even_subclassesB_derPhi,
					leftPerc_even_class1_derPhi, rightPerc_even_class1_derPhi,
					leftPerc_even_unseen1_derPhi, rightPerc_even_unseen1_derPhi,
					leftPerc_even_missclassified_derPhi, rightPerc_even_missclassified_derPhi]
	
	Tools.Plot_MediansOfFeatureMedias(princ_coord, [medians_even_subclassesA_derPhi, medians_even_subclassesB_derPhi], medians_even_class1_derPhi, mol['poreRadius_mean'], mol['poreRadius_std'], 
								   	 	medians_unseen = medians_even_unseen1_derPhi, medians_missclass = medians_even_missclassified_derPhi,  y_label = 'Interfacial inertia')
	
	##
	## derPhi plot (odd contributions) ##
	##
	percentiles = [leftPerc_odd_subclassesA_derPhi, rightPerc_odd_subclassesA_derPhi,
					leftPerc_odd_subclassesB_derPhi, rightPerc_odd_subclassesB_derPhi,
					leftPerc_odd_class1_derPhi, rightPerc_odd_class1_derPhi,
					leftPerc_odd_unseen1_derPhi, rightPerc_odd_unseen1_derPhi,
					leftPerc_odd_missclassified_derPhi, rightPerc_odd_missclassified_derPhi]
	
	Tools.Plot_MediansOfFeatureMedias(princ_coord,  [medians_odd_subclassesA_derPhi, medians_odd_subclassesB_derPhi], medians_odd_class1_derPhi, mol['poreRadius_mean'], mol['poreRadius_std'], 
								   		medians_unseen = medians_odd_unseen1_derPhi, medians_missclass = medians_odd_missclassified_derPhi, y_label = 'Interfacial conductivity')


'''
		Learn pp-specific thermostability properties		
'''
def PorePointLearning(features, class_0, class_1, fn_score, BALANCING = False, method = 'LogReg', kernel = None):
							
	# Get orientation	
	mol_info_fn = glob.glob('*_molInfo.xlsx')[0]
	wb = openpyxl.load_workbook(mol_info_fn)	 
	source = wb['geometry_information']
	ind_ppAxis	= source.cell(row = 2, column = 1).value 
	axis_ori = np.sign(ind_ppAxis) 
	ind_ppAxis = abs(ind_ppAxis) - 1
	
	# Load molecular information
	mol = Tools.LoadFile('mol')
	
	# Get principal pore axis coord
	princ_coord	= mol['porePoints'][:,ind_ppAxis]*axis_ori
	# nrOfRes = len(mol['resInfo'])
	nrOfPps = len(princ_coord)
	
	# Load variants	
	varInfo = Tools.LoadFile('varInfo')
	varInfo_type = Tools.GetColumn(varInfo, [1])

	data = []
	nrOfFeatures = len(features)
	for i in range(nrOfFeatures):		
		data.append(Tools.LoadFile(features[i]))

	# Get indices
	inds_class0 = Tools.GroupVariants(varInfo_type, class_0)
	inds_class1 = Tools.GroupVariants(varInfo_type, class_1)
	inds_class0_ = np.arange(0, len(inds_class0), 1)
	inds_class1_ = np.arange(len(inds_class0), len(inds_class0) + len(inds_class1), 1)
	
	# Count ..
	n_class_0 = len(inds_class0)
	n_class_1 = len(inds_class1)
	
	n_data = n_class_1 + n_class_0
	ratio = n_class_0/n_class_1
	
	print('\n The ratio of %s class instances to %s class instances is %.2f ' % (class_0, class_1, ratio))
	
	# Initialize
	f1_train = np.zeros((len(princ_coord), 5)) # mean, std, meadian, percentiles
	auc_train = np.zeros((len(princ_coord), 5)) 
	acc_train = np.zeros((len(princ_coord), 5))

	f1_test	= np.zeros((len(princ_coord), 5)) 
	auc_test = np.zeros((len(princ_coord), 5)) 
	acc_test = np.zeros((len(princ_coord), 5))
	
	features_importance	= np.zeros((len(princ_coord), nrOfFeatures, 5))
	
	X = np.zeros((n_data, nrOfFeatures)) # features of class_0 and class_1
	y = np.append(np.zeros(len(inds_class0)), np.ones(len(inds_class1))) # classes of class_0 and class_1
	
	print('Note that the smallest data set will contain approx %d data points!' % math.ceil(n_data/ModelParameters.K))

	# Initialize 
	prob_class_0_tested	= []
	
	for i in range(nrOfPps):
	
		# Initialize
		data_class_0 = []
		data_class_1 = []
		
		for j in range(nrOfFeatures):
			
			X[:,j] = np.append(data[j][i][inds_class0_], data[j][i][inds_class1_])  # the first len(inds_class0_) elements belong to class 0
				
			if (np.isnan(X[:,j]).any() == True): 
				exit('\n\n .. Exiting smoothly .. You have NaNs in your features! .. \n\n')
				
			## Append ##
			data_class_0.append(data[j][i][inds_class0_])	
			data_class_1.append(data[j][i][inds_class1_])	
					
		# Perform K-fold cross-validation Log Reg	
		# each array is of length (K - 1) * NUM_OF_TRAININGS 
		prob_class_0_tested_i, _, acc_auc_i, features_importance_i = Tools.KFoldsClassifier(X, y, method = method, kernel = kernel)
		
		## Append 
		# We calcualte the median summarizing (K - 1) * NUM_OF_TRAININGS training cycles
		prob_class_0_tested.append(np.median(prob_class_0_tested_i, axis = 0))			
		
		## Stats
		features_importance[i,:,0] = np.mean(features_importance_i, axis = 0)
		features_importance[i,:,1] = np.std(features_importance_i, axis = 0)
		features_importance[i,:,2] = np.median(features_importance_i, axis = 0)
		features_importance[i,:,3] = np.percentile(features_importance_i, q=ModelParameters.P_LEFT, axis = 0)
		features_importance[i,:,4] = np.percentile(features_importance_i, q=100-ModelParameters.P_LEFT, axis = 0)

		acc_auc_mean = np.mean(acc_auc_i, axis = 0)
		acc_auc_std = np.std(acc_auc_i, axis = 0)
		acc_auc_median = np.median(acc_auc_i, axis = 0)
		acc_auc_percentile_l = np.percentile(acc_auc_i, q=ModelParameters.P_LEFT, axis = 0)
		acc_auc_percentile_r = np.percentile(acc_auc_i, q=100-ModelParameters.P_LEFT, axis = 0)
		
		acc_train[i,:] = acc_auc_mean[0], acc_auc_std[0], acc_auc_median[0], acc_auc_percentile_l[0], acc_auc_percentile_r[0]
		acc_test[i,:] = acc_auc_mean[1], acc_auc_std[1], acc_auc_median[1], acc_auc_percentile_l[1], acc_auc_percentile_r[1]
		auc_train[i,:] = acc_auc_mean[2], acc_auc_std[2], acc_auc_median[2], acc_auc_percentile_l[2], acc_auc_percentile_r[2]
		auc_test[i,:] = acc_auc_mean[3], acc_auc_std[3], acc_auc_median[3], acc_auc_percentile_l[3], acc_auc_percentile_r[3]
		f1_train[i,:] = acc_auc_mean[4], acc_auc_std[4], acc_auc_median[4], acc_auc_percentile_l[4], acc_auc_percentile_r[4]
		f1_test[i,:] = acc_auc_mean[5], acc_auc_std[5], acc_auc_median[5], acc_auc_percentile_l[5], acc_auc_percentile_r[5]	
		
	# We have aggregated all the local class_0 probability scores in terms of the median.
	# That is, each residue is represented only by one value, i.e., its median class_0 probability
	# This is very simple and serves here as a proof of concept.
	# More clever way to aggreate may be considered
	
	# plt.hist(scores[inds_class0_], alpha = 0.2, color = "r")
	# plt.hist(scores[inds_class1_], alpha = 0.2, color = "b")
	# plt.show()
	
	## Store
	Tools.StoreFile(scores, fn_score[0] +  '_scores')	
	Tools.StoreFile(auc_train, fn_score[0] + '_auc_train')
	Tools.StoreFile(f1_train, fn_score[0] + '_f1_train')
	Tools.StoreFile(auc_test, fn_score[0] + '_auc_test')
	Tools.StoreFile(f1_test, fn_score[0] + '_f1_test')
	Tools.StoreFile(features_importance, fn_score[0] + '_featuresImportance')


"""
		Plot pp-specific learnings
"""	
def Plot_Learnings(fn_score, feat_labels):
	
	# Get orientation	
	mol_info_fn = glob.glob('*_molInfo.xlsx')[0]
	wb = openpyxl.load_workbook(mol_info_fn)	 
	source = wb['geometry_information']
	ind_ppAxis	= source.cell(row = 2, column = 1).value 
	axis_ori = np.sign(ind_ppAxis) 
	ind_ppAxis = abs(ind_ppAxis) - 1

	# Load molecular information
	mol = Tools.LoadFile('mol')
	# Get principal pore axis coord
	princ_coord	= mol['porePoints'][:,ind_ppAxis]*axis_ori
	
	auc_train = Tools.LoadFile(fn_score[0] + '_auc_train')
	f1_train = Tools.LoadFile(fn_score[0] + '_f1_train')
		
	auc_test = Tools.LoadFile(fn_score[0] + '_auc_test')
	f1_test = Tools.LoadFile(fn_score[0] + '_f1_test')

	# Plot features importance
	features_importance = Tools.LoadFile(fn_score[0] + '_featuresImportance')
	Tools.Plot_FeaturesImportance(princ_coord, features_importance, mol['poreRadius_mean'], feat_labels)
		
	# Algorithm performance (across pore points)
	fig, ax1 = plt.subplots()
	ax1.plot(princ_coord, auc_train[:,2], 'b-', label = 'AUC (train)')
	ax1.plot(princ_coord, f1_train[:,2], 'b--', label = 'f1 (train)')
	ax1.fill_between(princ_coord, auc_train[:,3], auc_train[:,4], color='b', linewidth=0, alpha = 0.25)
	ax1.plot(princ_coord, auc_test[:,2], 'g-', label = 'AUC (validation)')
	ax1.plot(princ_coord, f1_test[:,2], 'g--', label = 'f1 (validation)')
	ax1.fill_between(princ_coord, auc_test[:,3], auc_test[:,4], color='g', linewidth=0, alpha = 0.25)
	ax1.axhline(y = 0.5, alpha = 0.1, color = 'k')
	ax1.axhline(y = 0.8, alpha = 0.1, color = 'k')
	ax1.axhline(y = 0.7, alpha = 0.1, color = 'k')
	ax1.axhline(y = 0.6, alpha = 0.1, color = 'k')
	ax1.axhline(y = 0.9, alpha = 0.1, color = 'k')
	ax1.set_ylabel(r'Classifier performance ')
	# ax1.legend(fontsize = 40, loc = 'lower left')
	ax1.set_ylim([0,1])
	# ax1.set_xlabel(r'Pore point ($\perp$-coord.)')
	
	ax2 = ax1.twinx()
	ax2.plot(princ_coord, mol['poreRadius_mean'] / max(mol['poreRadius_mean']), 'k', label = r'$R/\mathrm{max}\{ R\}$' )
	ax2.fill_between(princ_coord, (mol['poreRadius_mean'] - mol['poreRadius_std'])  / max(mol['poreRadius_mean']), (mol['poreRadius_mean'] + mol['poreRadius_std'])  / max(mol['poreRadius_mean']), alpha = 0.1, color = 'k')
	ax2.set_xlim([min(princ_coord), max(princ_coord)])
	ax2.set_ylabel(r'Pore radius $R/\mathrm{max}\{ R\}$')
	# ax2.legend(fontsize = 40, loc = 'lower right')
	ax2.set_xticks([])
	# ax2.set_xticks([-14, -3, 8])
	# ax2.set_xticklabels(['AG', 'CC', 'SF'],  fontsize=70)
	plt.show()


'''
		Learn from aggregate pp-specific learnings
'''
def EnsembleLearning(	features, 

						class_0, 
						class_1, 
					
						method = 'LogReg', kernel = None):
	
	
	# Get orientation	
	mol_info_fn = glob.glob('*_molInfo.xlsx')[0]
	wb = openpyxl.load_workbook(mol_info_fn)	 
	source = wb['geometry_information']
	ind_ppAxis	= source.cell(row = 2, column = 1).value 
	ind_ppAxis = abs(ind_ppAxis) - 1
	source = wb['variants_information']
	offsetID = source.cell(row = 6, column = 1).value # do not forget the offset!
	offsetVal = source.cell(row = 4, column = 1).value # do not forget the offset!
	
	# Load molecular information
	mol = Tools.LoadFile('mol')	
	# Get principal pore axis coord
	# nrOfRes = len(mol['resInfo'])
	
	# Load variants	
	varInfo = Tools.LoadFile('varInfo')
	varInfo_resID = Tools.GetColumn(varInfo, [0])
	varInfo_type = Tools.GetColumn(varInfo, [1])
	varInfo_resNAME = Tools.GetColumn(varInfo, [2])
	
	# Load scores
	scores = []
	nrOfScores = len(features)
	for i in range(nrOfScores):
		scores.append(Tools.LoadFile(features[i] + '_scores'))

	data = []
	for i in range(nrOfScores):
		data.append(scores[i])			
	nrOfDataEntries = len(data)

	# Get indices
	inds_class0 = Tools.GroupVariants(varInfo_type, class_0)
	inds_class1 = Tools.GroupVariants(varInfo_type, class_1)
	inds_class0_ = np.arange(0, len(inds_class0), 1)
	inds_class1_ = np.arange(len(inds_class0), len(inds_class0) + len(inds_class1), 1)	

	# Count ..
	n_class_0 = len(inds_class0)
	n_class_1 = len(inds_class1)
	n_data = n_class_1 + n_class_0
	ratio = n_class_0/n_data	
		
	print('\n The ratio of %s class instances to %s class instances is %.2f ' % (class_0, class_1, ratio))
	print('\n Number of cases considered:  %s ' % (n_data))
	
	# Initialize
	scores = np.zeros((n_data, nrOfDataEntries))
			
	for k in range(nrOfDataEntries):
			
		scores[:,k] = np.append(data[k][inds_class0_], data[k][inds_class1_])
			
		if (np.isnan(scores[:,k]).any() == True): 
			exit('\n\n .. Exiting smoothly .. You have NaNs in your scores! .. \n\n')
			
	# Form target
	target = np.append(np.zeros(len(inds_class0)), np.ones(len(inds_class1)))
	# KFold Machine Learning procedure	
	res = Tools.KFoldsClassifier(scores, target, method = method, kernel = kernel)
	from sklearn.metrics import precision_recall_curve, f1_score

	# Obtain optimal threshold (f1)
	# Based on the median (aggregated) class_0 probabitlies! Note that this our global probability measure 
	# Also, note that the ground truth is 1 - target
	prob_class_0 = np.median(res[0], axis = 0) # the larger, the more likely to belong to class_0
	# Perform bootstrapping
	best_thresholds, _ = Tools.BootstrapThreshold(1 - target, prob_class_0, n_iterations=1000)
	# Calculate 95% confidence interval for the threshold
	lower_ci, upper_ci = np.percentile(best_thresholds, [2.5, 97.5])

	# Print results
	print(f"Best Threshold (mean): {np.mean(best_thresholds)}")
	print(f"Best Threshold (median): {np.median(best_thresholds)}")
	print(f"Standard Deviation of Thresholds: {np.std(best_thresholds)}")
	print(f"95% Confidence Interval for Threshold: ({lower_ci}, {upper_ci})")

	""" 
	print('\n Classification (class 0): ', np.median(res[0], axis = 0)[0:n_class_0])
	print('\n Classification (class 0): ', np.percentile(res[0], q=ModelParameters.P_LEFT, axis = 0)[0:n_class_0])
	print('\n Classification (class 0): ', np.percentile(res[0], q=100 - ModelParameters.P_LEFT, axis = 0)[0:n_class_0])
	print('\n ', varInfo_resID[inds_class_0])
	"""

	# Create the final plot with a linear threshold retrieved from f1-score maximization
	if (class_0 == "GoF/LoF"):

		# create the names
		class0_names = Tools.ResidueNameAndID(varInfo_resNAME[inds_class0], varInfo_resID[inds_class0], offsetID, offsetVal) 
		class1_names = Tools.ResidueNameAndID(varInfo_resNAME[inds_class1], varInfo_resID[inds_class1], offsetID, offsetVal) 
		# append
		resIDs = np.append(varInfo_resID[inds_class0], varInfo_resID[inds_class1])
		# get the indices
		inds_PEPD = Tools.GroupVariants(varInfo_type, "PEPD")
		inds_SFN = Tools.GroupVariants(varInfo_type, "SFN")
		inds_IEM = Tools.GroupVariants(varInfo_type, "IEM")
		inds_LoF = Tools.GroupVariants(varInfo_type, "LoF")
		LoF_resID = varInfo_resID[inds_LoF]
		PEPD_resID = varInfo_resID[inds_PEPD]
		SFN_resID = varInfo_resID[inds_SFN]
		IEM_resID = varInfo_resID[inds_IEM]
		indsAppend_LoF = np.where(np.isin(resIDs, LoF_resID))[0] # indices where LoF_ID appear in class0_ID
		indsAppend_PEPD = np.where(np.isin(resIDs, PEPD_resID))[0] # indices where GoF_ID appear in class0_ID
		indsAppend_SFN = np.where(np.isin(resIDs, SFN_resID))[0] # indices where GoF_ID appear in class0_ID
		indsAppend_IEM = np.where(np.isin(resIDs, IEM_resID))[0] # indices where GoF_ID appear in class0_ID
		# append, calcualte the thershold, and plot
		indsAppend_GoF = [indsAppend_PEPD, indsAppend_SFN, indsAppend_IEM]
		optThres = np.mean(best_thresholds) 
		optThres_left = optThres - 2 * np.std(best_thresholds)
		optThres_right = optThres + 2 * np.std(best_thresholds) 
		# box plots: lines indicated the medians
		Tools.Plot_MutationScores(res[0], np.append(class0_names, class1_names), n_class_0, n_class_1, indsAppend_LoF, indsAppend_GoF, [optThres, optThres_left, optThres_right])

	# Report key-findigns
	# Print the medians of the class 0 probabilities:
	# print('\n Classification (class 0) status: ', np.median(res[0], axis = 0)[0:n_class_0])
	# print('\n ', varInfo_resID[inds_class0])
	# print('\n Classification (class 1) status: ', np.median(res[0], axis = 0)[n_class_0:n_data])
	# print('\n ', varInfo_resID[inds_class1])
				
	print('\n\n AUC median scoring: ')	
	print('train: %2f' % np.round(np.median(Tools.GetColumn(res[2], [2])),2))	
	print('Percentiles (train):', np.round(np.percentile(Tools.GetColumn(res[2], [2]),q=5),2), np.round(np.percentile(Tools.GetColumn(res[2], [2]),q=95),2))
	print('test: %2f' % np.round(np.median(Tools.GetColumn(res[2], [3])),2))	
	print('Percentiles (test):', np.round(np.percentile(Tools.GetColumn(res[2], [3]),q=5),2), np.round(np.percentile(Tools.GetColumn(res[2], [3]),q=95),2))

	print('\n\n F1 median scoring: ')
	print('train: %2f' % np.round(np.median(Tools.GetColumn(res[2], [4])),2))	
	print('Percentiles (train):', np.round(np.percentile(Tools.GetColumn(res[2], [4]),q=5),2), np.round(np.percentile(Tools.GetColumn(res[2], [4]),q=95),2))											
	print('test: %2f' % np.round(np.median(Tools.GetColumn(res[2], [5])),2))	
	print('Percentiles (test):', np.round(np.percentile(Tools.GetColumn(res[2], [5]),q=5),2), np.round(np.percentile(Tools.GetColumn(res[2], [5]),q=95),2))														

	# AUC
	plt.title("AUC-score")
	plt.hist(Tools.GetColumn(res[2], [2]), alpha = 0.2, color = 'blue', label = r'Training (GoF$\cup$LoF-vs.-Neutr.$\cup$Benign)', edgecolor='black', bins = 30) # 
	plt.hist(Tools.GetColumn(res[2], [3]), alpha = 0.2, color = 'green',  label = r'Validation (GoF$\cup$LoF-vs.-Neutr.$\cup$Benign)', edgecolor='black', bins = 30)
	# plt.hist(Tools.GetColumn(res[-1], [1]), color = 'red', alpha = 0.2, label = r'Test ("unseen" data)', edgecolor='black')
	plt.xlim([0,1])
	plt.xticks([0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.76, 0.9])
	plt.axvline(x = 0.76, linestyle = "--", color = "k", linewidth = 5, alpha = 0.5)
	plt.legend(fontsize = 40, loc = 'upper left')
	plt.xlabel('Performance (AUC)', fontsize = 50)
	plt.ylabel('Freq.',fontsize = 45)
	plt.ylim([0,15])
	plt.show()

	# f1-score
	plt.title("f1-score")
	plt.hist(Tools.GetColumn(res[2], [5]), alpha = 0.2, color = 'blue', label = r'Training (GoF$\cup$LoF-vs.-Neutr.)', edgecolor='black', bins = 30)
	plt.hist(Tools.GetColumn(res[2], [4]), alpha = 0.2, color = 'green', label = r'Validation (GoF$\cup$LoF-vs.-Neutr.)', edgecolor='black', bins = 30)
	# plt.hist(Tools.GetColumn(res[-1], [2]), color = 'red', alpha = 0.2,  label = r'Test ("unseen" data)', edgecolor='black')
	plt.xlim([0,1])
	plt.xticks([0,0.1,0.2,0.3,0.4,0.5,0.6,0.78,0.9])
	plt.axvline(x = 0.78, linestyle = "--", color = "k", linewidth = 5, alpha = 0.5)
	plt.legend(fontsize = 40, loc = 'upper left')
	plt.xlabel('Performance (f1-score)', fontsize = 45)
	plt.ylabel('Freq.',fontsize = 45)
	plt.ylim([0,50])
	plt.show()


'''
		Collect observables 
'''	
def CollectObservables():
	
	mol = Tools.LoadFile('mol')
		
	## Init 
	nrOfPps = len(mol['poreRadius_mean'])
	scales = []
	nrOfAtoms = []
	statMod = []
	dists = []
	hydrMoments = []
	entropies = []
	topologies = []
	exponentsDomains = []	
	order_odd_mom = np.arange(1,2*ModelParameters.N_HYDR, 2).astype(int)
	order_even_mom = np.arange(0,2*ModelParameters.N_HYDR, 2).astype(int)
	
	cumulativeAtomProbs = 0
	cumulativeResidueProbs = 0
	
	for i in range(nrOfPps):    
		
		# Calculate distances from the current pore point
		porePoint = mol['porePoints'][i,:]

		# print(' Pore point index %d (%.1f)' % (i, porePoint[2]))

		l_atom = Tools.EuclNorm(mol['coords_atom'], porePoint)
		l_res = Tools.EuclNorm(mol['coords_res'], porePoint)
				
		# Get molecular size
		maxPoreDist = max(l_atom)
		minPoreDist = min(l_atom)
		if (minPoreDist < mol['poreRadius_mean'][i]):
			print(minPoreDist, mol['poreRadius_mean'][i], mol['poreRadius_std'][i])
			print('\n\n .. Warning .. The pore radius is too big to be true! \n\n')
		ind_max = np.where(maxPoreDist == l_atom)[0]
		outterPoreRad = maxPoreDist + mol['vdWrad'][ind_max][0]
		
		## Append ##
		dists.append([l_atom, l_res])
		
		# get shell boundaries (nr. of shells (scales): Parameters.N_SCALES)
		shell_boundaries = np.linspace(minPoreDist, outterPoreRad, num = ModelParameters.N_SCALES + 1).flatten()
		scales.append(shell_boundaries[1:len(shell_boundaries)])		
	
		# initilize arrays for shell calculations (much faster than whole sphere and/or list calculations ..)
		N_sphere = [0] * 3
		h_even_prev_sphere = [[0] * ModelParameters.N_HYDR] * 3
		h_odd_prev_sphere = [[np.zeros(3)] * ModelParameters.N_HYDR] * 3
		nrOfAtoms_sphere = []
		hydrEnerg_sphere = []
		inds_pre_sphere_set	= set([])
		
		for j in range(ModelParameters.N_SCALES):
		
			# get shell boundary size
			up_bound = scales[i][j]
			# Get inds of all components INSIDE the sphere by using a heaviside function
			inds_sphere	= np.where(l_atom < up_bound)[0]
			# Get inds of all components INSIDE the spherical shell 
			inds_sphere_set = set(inds_sphere)
			inds_shell = list(inds_pre_sphere_set ^ inds_sphere_set)
			# Update
			inds_pre_sphere_set	= inds_sphere_set
			
			## Calculate spherical shell properties	
			# Atom-level
			N_sphere[0] += len(inds_shell)
			sampled_coords = mol['coords_atom'][inds_shell]
			sampled_hvals = mol['hvals'][inds_shell]
			# phobic
			sampled_coords_pho = sampled_coords[np.where(sampled_hvals < 0)[0]]
			sampled_hvals_pho = sampled_hvals[np.where(sampled_hvals < 0)[0]]
			N_sphere[1] += len(sampled_hvals[np.where(sampled_hvals < 0)[0]])
			# philic
			sampled_coords_phi = sampled_coords[np.where(sampled_hvals >= 0)[0]]
			sampled_hvals_phi = sampled_hvals[np.where(sampled_hvals >= 0)[0]]
			N_sphere[2] += len(sampled_hvals[np.where(sampled_hvals >= 0)[0]])
			
			# Introduce local origin (INSIDE coords)
			vecs = sampled_coords - porePoint
			vecs_pho = sampled_coords_pho - porePoint
			vecs_phi = sampled_coords_phi - porePoint

			# Calculate sphere properties
			# .. for all the atoms
			h_even_sphere = Tools.EvenOrderHydrMom(vecs, sampled_hvals, order_even_mom) + np.asarray(h_even_prev_sphere[0])
			h_odd_sphere = Tools.OddOrderHydrMom(vecs, sampled_hvals, order_odd_mom) + np.asarray(h_odd_prev_sphere[0])
			# .. for the hydrophobic atoms only
			h_even_sphere_pho = Tools.EvenOrderHydrMom(vecs_pho, sampled_hvals_pho, order_even_mom) + np.asarray(h_even_prev_sphere[1])
			h_odd_sphere_pho = Tools.OddOrderHydrMom(vecs_pho, sampled_hvals_pho, order_odd_mom) + np.asarray(h_odd_prev_sphere[1])		
			# .. for the hydrophilic atoms only
			h_even_sphere_phi = Tools.EvenOrderHydrMom(vecs_phi, sampled_hvals_phi, order_even_mom) + np.asarray(h_even_prev_sphere[2])
			h_odd_sphere_phi = Tools.OddOrderHydrMom(vecs_phi, sampled_hvals_phi, order_odd_mom) + np.asarray(h_odd_prev_sphere[2])	
		
			## Append ## IF the total hydr density AND the total hydr first-order mom z-component are correct ##
			if ((N_sphere[0] == N_sphere[1] + N_sphere[2]) and abs(h_even_sphere[0] - (h_even_sphere_pho[0] + h_even_sphere_phi[0])) < ModelParameters.ACC and abs(h_odd_sphere[0][2] - (h_odd_sphere_pho[0][2] + h_odd_sphere_phi[0][2])) < ModelParameters.ACC):
				hydrEnerg_sphere.append([[h_even_sphere, h_odd_sphere], [h_even_sphere_pho, h_odd_sphere_pho], [h_even_sphere_phi, h_odd_sphere_phi]])
				nrOfAtoms_sphere.append([N_sphere[0], N_sphere[1], N_sphere[2]])
			else:
				exit('\n\n .. Exiting smoothly .. Check your hydr mom calculations .. \n\n')
			
			# update previous sphere content (even) 
			h_even_prev_sphere[0] = h_even_sphere
			h_even_prev_sphere[1] = h_even_sphere_pho
			h_even_prev_sphere[2] = h_even_sphere_phi
			# update previous sphere content (odd)
			h_odd_prev_sphere[0] = h_odd_sphere
			h_odd_prev_sphere[1] = h_odd_sphere_pho
			h_odd_prev_sphere[2] = h_odd_sphere_phi
			
		
		if (abs(N_sphere[0] - mol['total_nrOfAtoms']) < ModelParameters.ACC):
			# Retrieve the initial values for performing parameter estimation for the stat model .
			l = np.asarray(scales[i]) 				
			N = Tools.GetColumn(nrOfAtoms_sphere, [0])
			N_pho = Tools.GetColumn(nrOfAtoms_sphere, [1])
			N_phi = Tools.GetColumn(nrOfAtoms_sphere, [2])
			_ = Tools.StatModelParameters(l, N/max(N)) 
			_pho = Tools.StatModelParameters(l, N_pho/max(N_pho))
			_phi = Tools.StatModelParameters(l, N_phi/max(N_phi))
			statMod.append(	
							[
								[value for value in _.values()    if isinstance(value, (int, float)) or value is None], 
								[value for value in _pho.values() if isinstance(value, (int, float)) or value is None], 
								[value for value in _phi.values() if isinstance(value, (int, float)) or value is None]
							]
						)
						
			## Append ## 
			nrOfAtoms.append([N, N_pho, N_phi])  
			# Get the probabilities
			# Note that sum(p_unnorm) is the partition sum: p = p_unnorm/sum(p_unnorm) = (n/K)/(sum(n)/K) = n/sum(n)
			p_unnorm = Tools.GeomModel(l, [   _['K'] ,    _['inv_zeta'] ,    _['l_i'] ,    _['nu'] ],    _["modType"] ) /  _['K'] 
			p_pho_unnorm = Tools.GeomModel(l, [_pho['K'] , _pho['inv_zeta'] , _pho['l_i'] , _pho['nu'] ], _pho["modType"] ) / _pho['K']
			p_phi_unnorm = Tools.GeomModel(l, [_phi['K'] , _phi['inv_zeta'] , _phi['l_i'] , _phi['nu'] ], _phi["modType"] ) / _phi['K']
			# Radial order entropies
			S = Tools.qEntropy(p_unnorm/sum(p_unnorm), _['nu'] + 1)
			S_pho = Tools.qEntropy(p_pho_unnorm/sum(p_pho_unnorm), _pho['nu'] + 1)
			S_phi = Tools.qEntropy(p_phi_unnorm/sum(p_phi_unnorm), _phi['nu'] + 1)
			# Curvature (and corrsponding domains)
			curvature = Tools.GeomModel(l, [_['K'] , _['inv_zeta'] , _['l_i'] , _['nu'] ], _["modType"], 2) / _['K']
			ind_max_curv = Tools.Match(max(curvature), curvature)[0]
			ind_min_curv = Tools.Match(min(curvature), curvature)[0]
			# Cumulative probabilities (roughly determine the "phase" (i.e., PD vs VS) to which an atom belongs)
			p_res = Tools.GeomModel(l_res,  [_['K'] , _['inv_zeta'] , _['l_i'] , _['nu'] ], _["modType"] ) / _['K']
			p_atom = Tools.GeomModel(l_atom, [_['K'] , _['inv_zeta'] , _['l_i'] , _['nu'] ], _["modType"] ) / _['K']
		
		else:
			exit('\n .. Exiting smoothly .. Some atoms were NOT sampled! \n\n')
		
		
		# Collect hydr energies and entropy
		## Append ## 
		hydrMoments.append(hydrEnerg_sphere)
		## Append ##
		entropies.append([S, S_pho, S_phi])
		
		# i-th data collection for scaling of domain and topologies
		zeroCrossings_i = []
		exponentsDomain_i = []

		# Domains scaling and topology
		for order in order_odd_mom:

			# EVEN
			h_even = Tools.getHydrMom(hydrMoments, i, 'pathic', order - 1)
		
			# ODD
			h_odd_z = Tools.getHydrMom(hydrMoments, i, 'pathic', order, 'z')
				
			# sign-changes detection
			zeroCrossings_i.append([	
										Tools.ZeroCrossings(l, h_even),
										Tools.ZeroCrossings(l, h_odd_z),								
								   ])

			# power-law envelope approximation
			exponentsDomain_i.append([	
										Tools.ScalingBehavior(l, abs(h_even), np.array([l[ind_max_curv], _['l_i'], l[ind_min_curv]])),
										Tools.ScalingBehavior(l, abs(h_odd_z), np.array([l[ind_max_curv], _['l_i'], l[ind_min_curv]])),
									])
		
			
											
		# Increment ..
		cumulativeAtomProbs += p_atom
		cumulativeResidueProbs += p_res
		
		## Append ##
		topologies.append(zeroCrossings_i)
		## Append ##
		exponentsDomains.append(exponentsDomain_i)
	
	# Store ..
	Tools.StoreFile(cumulativeAtomProbs/nrOfPps, 'atomProbs')
	Tools.StoreFile(cumulativeResidueProbs/nrOfPps, 'resProbs')
	Tools.StoreFile(scales, 'scales')
	Tools.StoreFile(dists, 'dists')
	Tools.StoreFile(nrOfAtoms, 'nrOfAtoms')
	Tools.StoreFile(statMod, 'statMod')
	Tools.StoreFile(hydrMoments, 'hydrMoments')
	Tools.StoreFile(entropies, 'entropies')
	Tools.StoreFile(topologies, 'topologies')
	Tools.StoreFile(exponentsDomains, 'exponentsDomains')


'''
		Information profile
'''
def InformationProfile(poreAxisLimits = ['min','max'], orderOfMom = 1, PLOT = False, statsType = 'median'):
	
	exponentsDomains = Tools.LoadFile('exponentsDomains')
	statMod = Tools.LoadFile('statMod')
	scales = Tools.LoadFile('scales')
	hydrMoments = Tools.LoadFile('hydrMoments')
	nrOfAtoms = Tools.LoadFile('nrOfAtoms')
	entropies = Tools.LoadFile('entropies')
	cumulativeAtomProbs = Tools.LoadFile('atomProbs')
	mol = Tools.LoadFile('mol')
	entropy = Tools.GetColumn(entropies, [0])
	
	# Get orientation	
	mol_info_fn = glob.glob('*_molInfo.xlsx')[0]
	wb = openpyxl.load_workbook(mol_info_fn)	 
	source = wb['geometry_information']
	ind_ppAxis	= source.cell(row = 2, column = 1).value 
	axis_ori = np.sign(ind_ppAxis) 
	ind_ppAxis = abs(ind_ppAxis) - 1
		
	# Get principal pore axis coord
	princ_coord	= mol['porePoints'][:,ind_ppAxis]*axis_ori
	nrOfPps = len(princ_coord)

	# Modeling informatio	
	A = Tools.GetColumn(statMod, [0, 0])	
	a = Tools.GetColumn(statMod, [0, 2])
	l_i = Tools.GetColumn(statMod, [0, 4])
	lag = Tools.GetColumn(statMod, [0, 8])
	asy = Tools.GetColumn(statMod, [0, 10])
	nu 	= Tools.GetColumn(statMod, [0, 6])
	nu_unc 	= Tools.GetColumn(statMod, [0, 7])
	
	MAFE = Tools.GetColumn(statMod, [0, 12])
	pval = Tools.GetColumn(statMod, [0, 13])
	modType = Tools.GetColumn(statMod, [0, 15])
	inds_gomp = np.where(modType == 3)[0]

	invxi = nu*a 
	invxi[inds_gomp] = a[inds_gomp]
	xi = 1./invxi 
	# introduce \zeta
	zeta = 1/a
	zeta[inds_gomp] =  1/float('inf')

	data_n = []
	data_n_model = []
	data_m0 = [] 
	data_m0_pho = [] 
	data_m0_phi = [] 
	data_expModel = [] 

	data_h_z_pos = []
	data_h_z_ES = []
	data_h_z_neg = []
	data_h_z_IS = []
	data_h_z_standard = []
	data_phi_1 = []
		
	l_max = np.zeros(nrOfPps)
	l_max_curv = np.zeros(nrOfPps)
	l_min = np.zeros(nrOfPps)
	l_min_curv = np.zeros(nrOfPps)
	inds_l_i = np.zeros(nrOfPps)
	fracDim_i = np.zeros(nrOfPps)
	fracDim_max = np.zeros(nrOfPps)
	fracDim_i_exp = np.zeros(nrOfPps)
	T = np.zeros(nrOfPps)
	
	gamma_exp_PD = np.zeros(nrOfPps)
	PC_gamma_exp_PD = np.zeros(nrOfPps)
	gamma_exp_VSDs = np.zeros(nrOfPps)
	PC_gamma_exp_VSDs = np.zeros(nrOfPps)
	l_half_emp = np.zeros(nrOfPps)

	for i in range(nrOfPps):  
		
		# radii
		l = scales[i]
		l_min[i], l_max[i] = min(l), max(l)
		ind_l_i = np.argmin(abs(l - l_i[i]))
		inds_l_i[i] = ind_l_i

		# nr of atoms and unit-mass-fractal dimension at l_i
		n = nrOfAtoms[i][0]
		n_model = Tools.GeomModel(l, [A[i],a[i],l_i[i],nu[i]], modType[i])*max(n)
		fracDim_i[i] = Tools.GeomModel(l, [A[i],a[i],l_i[i],nu[i]], modType[i], der = 'log0')[ind_l_i]
		fracDim_max[i] = max(Tools.GeomModel(l, [A[i],a[i],l_i[i],nu[i]], modType[i], der = 'log0'))

		# l_* empirical
		abs_diff = np.abs(n/max(n) - 0.5)
		ind_l_half = np.argmin(abs_diff)
		l_half_emp[i] = l[ind_l_half]
		
		# normalization and calculation of packing probabilties
		n_norm	= n_model / A[i] 
		Z_p	= sum(n_norm)
		p = n_norm / Z_p 

		# curvature of n
		curvature = Tools.GeomModel(l, [A[i],a[i],l_i[i],nu[i]], modType[i], der=2) / A[i]
		ind_max_curv = Tools.Match(max(curvature), curvature)[0]
		ind_min_curv = Tools.Match(min(curvature), curvature)[0]
		l_max_curv[i] = l[ind_max_curv]
		l_min_curv[i] = l[ind_min_curv]

		# experimental fractal dimension at l_i
		coeffs, _ = np.polyfit(np.log(l[ind_max_curv:ind_min_curv]), np.log(n[ind_max_curv:ind_min_curv]), deg=1, cov=True)
		fracDim_i_exp[i] = coeffs[0]
		
		"""
		linModel_exp = coeffs[0] * np.log(l[ind_max_curv:ind_min_curv]) + coeffs[1]
		linModel_theory = fracDim_i[i] * np.log(l[ind_max_curv:ind_min_curv]) + coeffs[1]
		plt.plot(np.log(l[ind_max_curv:ind_min_curv]), np.log(n[ind_max_curv:ind_min_curv]), "bo", alpha = 0.2)
		plt.plot(np.log(l[ind_max_curv:ind_min_curv]), linModel_theory, "r--", linewidth = 4, alpha = 0.5)
		plt.plot(np.log(l[ind_max_curv:ind_min_curv]), linModel_exp, "g--", linewidth = 4, alpha = 0.5)
		plt.show()
		"""
		
		# hydropathic energy magnitude
		m0_abs = abs(Tools.getHydrMom(hydrMoments, i, 'pathic', 0)) / (n) 
		m0_pho = Tools.getHydrMom(hydrMoments, i, 'phobic', 0) / (n) 
		m0_phi = Tools.getHydrMom(hydrMoments, i, 'philic', 0) / (n) 

		h1_pho = Tools.getHydrMom(hydrMoments, i, 'phobic', 1) + ModelParameters.ZERO # add ZERO to avoid zeros
		h1_phi = Tools.getHydrMom(hydrMoments, i, 'philic', 1) + ModelParameters.ZERO

		# Check if h1 is decomposable (typically not)
		h_plus, h_minus = Tools.Decompose(h1_pho, h1_phi, ind_max_curv)

		# Define phi (for j=1) for increasing l (log of comp susc)
		phi_1 = np.log(abs(h_plus / h_minus))

		# Exponential model fit on hydr energy magnitude
		arg = l/xi[i]
		arg_i = l_i[i]/xi[i]
		expModelFit = scipy.optimize.curve_fit(Tools.ExpFunc, arg, m0_abs)
		expModel = expModelFit[0] * np.exp(- arg) 

		# Physical temperature calculation
		q = nu[i] + 1. # note that q > 1
		epsilon	= np.exp(- arg) 	
		epsilon_i = np.exp(- arg_i) 	
		Z_pi = sum((p**q)) # Escort distribution partition sum
		U = sum((p**q) * epsilon) / Z_pi # "internal" energy constraint
		T[i] = ( ( 1 - (1 - q)*U/epsilon_i ) / Z_pi ) * epsilon_i # measured in \Theta

		# Key thermostability obs (first-order hydr dip moment)
		h_z = Tools.getHydrMom(hydrMoments, i, 'pathic', orderOfMom, 'z')
		h_z_pho = Tools.getHydrMom(hydrMoments, i, 'phobic', orderOfMom, 'z') + ModelParameters.ZERO
		h_z_phi = Tools.getHydrMom(hydrMoments, i, 'philic', orderOfMom, 'z') + ModelParameters.ZERO
		
		# Define eff susc and extract coarse-grained scaling exponents
		susc = np.abs(h_z_phi/h_z_pho)
		coeffs_PD, _ = np.polyfit(np.log(l[ind_max_curv:ind_l_i]), np.log(susc[ind_max_curv:ind_l_i]), deg=1, cov=True)
		PC_PD = stats.pearsonr(np.log(susc[ind_max_curv:ind_l_i]), coeffs_PD[0]*np.log(l[ind_max_curv:ind_l_i]) + coeffs_PD[1])[0]
		coeffs_VSDs, _ = np.polyfit(np.log(l[ind_l_i:ind_min_curv]), np.log(susc[ind_l_i:ind_min_curv]), deg=1, cov=True)
		PC_VSDs = stats.pearsonr(np.log(susc[ind_l_i:ind_min_curv]), coeffs_VSDs[0]*np.log(l[ind_l_i:ind_min_curv]) + coeffs_VSDs[1])[0]		

		gamma_exp_PD[i] = coeffs_PD[0]
		PC_gamma_exp_PD[i] = PC_PD
		gamma_exp_VSDs[i] = coeffs_VSDs[0]
		PC_gamma_exp_VSDs[i] = PC_VSDs

		h_z_standarized = Tools.Standarization(h_z, l, ind_l_i)
		h_z_norm = h_z / abs(h_z[ind_l_i])

		# Separate the positive from the negative branch
		if (h_z[ind_l_i] > 0):
			data_h_z_pos.append(h_z_norm)
		else:
			data_h_z_neg.append(h_z_norm)	

		# Separate traces based on which side of the pore we are
		if (princ_coord[i] > 0):
			data_h_z_ES.append(h_z_norm)
		else:
			data_h_z_IS.append(h_z_norm)

		data_h_z_standard.append(h_z_standarized)
		data_n.append(n)   
		data_n_model.append(n_model)  
		data_m0.append(m0_abs) 
		data_m0_pho.append(m0_pho) 
		data_m0_phi.append(m0_phi) 
		data_expModel.append(expModel) 
		data_phi_1.append(phi_1)
	

	# get exponents different domains
	# PD
	expPre = Tools.ScalingInfo(exponentsDomains, orderOfhydrMom = orderOfMom, component = 'z', dataType =  'powerLaw_pre', infoBlock = 'exponent')
	expPre_unc = Tools.ScalingInfo(exponentsDomains, orderOfhydrMom = orderOfMom, component = 'z', dataType =  'powerLaw_pre', infoBlock = 'unc')[::2]
	PC_pre = Tools.ScalingInfo(exponentsDomains, orderOfhydrMom = orderOfMom, component = 'z', dataType =  'powerLaw_pre', infoBlock = 'PC')
	# VSD (infl2)
	expInfl2 = Tools.ScalingInfo(exponentsDomains, orderOfhydrMom = orderOfMom, component = 'z', dataType =  'powerLaw_infl2', infoBlock = 'exponent')
	expInfl2_unc = Tools.ScalingInfo(exponentsDomains, orderOfhydrMom = orderOfMom, component = 'z', dataType =  'powerLaw_infl2', infoBlock = 'unc')[::2]
	PC_infl2 = Tools.ScalingInfo(exponentsDomains, orderOfhydrMom = orderOfMom, component = 'z', dataType =  'powerLaw_infl2', infoBlock = 'PC')   

	# PLot different views according to our prob coloring scheme
	# .. Use the command spectrum b, blue_red in pymol on the _probBFactors.pdb molecule 
	Tools.Plot_MoleculeViews(ind_ppAxis, axis_ori, mol['coords_atom'], cumulativeAtomProbs, l_i, PDB)
	# Sanity test!
	Tools.Plot_CheckTraces(princ_coord, l_min, l_max, l_i, l_half_emp, zeta, xi, lag, asy, l_max_curv, l_min_curv, PDB)
	
	if (PLOT == True):
		
		Tools.Plot_CumulativeAtomNumber(data_n, data_n_model, min(inds_l_i), max(inds_l_i), [MAFE, pval, fracDim_i, fracDim_i_exp, fracDim_max], statsType = statsType)

		# within some appropriate range 
		if (poreAxisLimits[0] == 'min'):
			limit_l = min(princ_coord)
		else:
			limit_l = poreAxisLimits[0]
		if (poreAxisLimits[1] == 'max'):
			limit_r = max(princ_coord)
		else:
			limit_r = poreAxisLimits[1]

		indices = np.where((np.array(princ_coord) >= limit_l) & (np.array(princ_coord) <= limit_r))[0]

		x = zeta[indices]
		y = xi[indices]
		z = entropy[indices] 
		c = princ_coord[indices]
		Tools.Plot_3DShape(x, y, z, c, r'$\zeta$ [\AA]', r'$\xi$ [\AA]', r'$\mathcal{S}$ (norm.)', r'$\perp$-coord.')
	
		x = nu[indices]
		y = l_max[indices] - l_min[indices]
		z = entropy[indices]
		c = princ_coord[indices]
		Tools.Plot_3DShape(x, y, z, c, r'$\nu$', r'$s$ [\AA]', r'$\mathcal{S}$ (norm.)', r'$\perp$-coord.')
		
		Tools.Plot_AtomicHydropathicEnergy(data_m0, data_m0_pho, data_m0_phi, data_expModel, min(inds_l_i), max(inds_l_i), [zeta, xi], statsType = statsType)
		Tools.Plot_DipoleMoment(data_h_z_pos, data_h_z_IS, data_h_z_neg, data_h_z_ES, min(inds_l_i), max(inds_l_i), [expPre, expInfl2], [PC_pre, PC_infl2], statsType = statsType, x_min = -20, x_max = 25)
		Tools.Plot_ExponentsTraces(princ_coord, expPre, expPre_unc, PC_pre, expInfl2, expInfl2_unc, PC_infl2, nu, nu_unc, mol['poreRadius_mean'], mol['poreRadius_std'])
	
	## Append and Store
	# Data that are useful for summarizing channel subtype features
	SummaryInfo = []
	SummaryInfo.append(princ_coord)  				
	SummaryInfo.append(mol["poreRadius_mean"])		
	SummaryInfo.append(A)							
	SummaryInfo.append(xi)								
	SummaryInfo.append(zeta)
	SummaryInfo.append(l_i)
	SummaryInfo.append(inds_l_i)
	SummaryInfo.append(l_max - l_min)
	SummaryInfo.append(MAFE)
	SummaryInfo.append(pval)
	SummaryInfo.append(entropy)
	SummaryInfo.append(T)
	SummaryInfo.append(np.median(data_n, axis=0))
	SummaryInfo.append(np.median(data_n_model, axis=0))
	SummaryInfo.append([fracDim_i, fracDim_i_exp, fracDim_max])
	SummaryInfo.append(np.median(data_m0, axis=0))
	SummaryInfo.append(np.median(data_expModel, axis=0))
	SummaryInfo.append(np.median(data_m0_pho, axis=0))
	SummaryInfo.append(np.median(data_m0_phi, axis=0))

	SummaryInfo.append(np.median(data_h_z_pos, axis=0))
	SummaryInfo.append(np.median(data_h_z_neg, axis=0))

	SummaryInfo.append(np.median(data_h_z_standard, axis=0))

	SummaryInfo.append([expPre, expInfl2])
	SummaryInfo.append([PC_pre, PC_infl2])

	SummaryInfo.append([gamma_exp_PD, PC_gamma_exp_PD, gamma_exp_VSDs, PC_gamma_exp_VSDs]) # effective susceptibility exponents

	SummaryInfo.append(np.median(data_h_z_ES, axis=0))
	SummaryInfo.append(np.median(data_h_z_IS, axis=0))
	
	SummaryInfo.append([
						np.median(data_phi_1, axis=0), 
					 	np.min(data_phi_1, axis=0), 
						np.max(data_phi_1, axis=0)
						])

	Tools.StoreFile(SummaryInfo, "SummaryInfo")
	
	
"""
		Summary of thermostability profile
"""
def Summary(subtype = [], poreAxisLimits = ['min', 'max'], orderOfMom = 1):

	# Get the current directory
	current_directory = os.getcwd()

	# List all files in the current directory
	files = os.listdir(current_directory)

	# Filter files to include only those ending with '.txt'
	summaryFiles = [file for file in files if file.endswith('.txt')]

	# Count the number of .txt files
	nrOfSummaryFiles = len(summaryFiles)

	print('\n .. Summarizing %d channel configurations of %s' % (len(summaryFiles), subtype))
	
	for file in summaryFiles:
		print(file[:4])
	
	# exit()

	# Load data
	princ_coords = []
	pore_radii = []	
	As = []
	xis = []
	zetas = []
	l_is = []
	ind_l_is = []
	sizes = []
	MAFEs = []
	pvals = []
	entropies = []
	Ts = []

	ns = []
	n_models = []

	fracDims_i = []
	fracDims_i_exper = []
	fracDims_max = []

	m0s = []

	m0s_pho = []
	m0s_phi = []
	m0_models = []
	
	data_h_z_poss = []
	data_h_z_negs = []
	data_h_z_standards = []

	expPres = []
	expInfl2s = []
	
	PC_expPres = []
	PC_expInfl2s = []

	gamma_exps_PD = [] 
	gamma_exps_VSDs = [] 

	PC_gamma_exps_PD = []
	PC_gamma_exps_VSDs = []

	data_h_z_ESs = []
	data_h_z_ISs = []

	avNrOfPps = 0

	for i in range(nrOfSummaryFiles):

		print(files[i])
		
		fn = summaryFiles[i]
		data = Tools.LoadFile(fn, 0)
		
		princ_coords.extend(data[0]) 
		pore_radii.extend(data[1])
		
		As.extend(data[2])

		xis.extend(data[3])
		zetas.extend(data[4])
		l_is.extend(data[5])
		ind_l_is.extend(data[6])
		sizes.extend(data[7])
		MAFEs.extend(data[8])
		pvals.extend(data[9])
		entropies.extend(data[10])
		Ts.extend(data[11])
		ns.extend(data[12])
		n_models.extend(data[13])
		fracDims_i.extend(data[14][0])
		fracDims_i_exper.extend(data[14][1])
		fracDims_max.extend(data[14][2])
		m0s.extend(data[15])
		m0_models.extend(data[16])
		m0s_pho.extend(data[17])
		m0s_phi.extend(data[18])

		data_h_z_poss.extend(data[19])
		data_h_z_negs.extend(data[20])
		data_h_z_standards.extend(data[21])

		expPres.extend(data[22][0])
		expInfl2s.extend(data[22][1])

		PC_expPres.extend(data[23][0])
		PC_expInfl2s.extend(data[23][1])

		gamma_exps_PD.extend(data[24][0])
		PC_gamma_exps_PD.extend(data[24][1])
		gamma_exps_VSDs.extend(data[24][2])
		PC_gamma_exps_VSDs.extend(data[24][3])

		data_h_z_ESs.extend(data[25])
		data_h_z_ISs.extend(data[26])

		avNrOfPps += len(data[0])

	'''
			Entropy plots: defomation forces and nonextensivity
	'''
	
	# within some appropriate range 
	if (poreAxisLimits[0] == 'min'):
		limit_l = min(princ_coords)
	else:
		limit_l = poreAxisLimits[0]
	if (poreAxisLimits[1] == 'max'):
		limit_r = max(princ_coords)
	else:
		limit_r = poreAxisLimits[1]

	indices = np.where((np.array(princ_coords) > limit_l) & (np.array(princ_coords) < limit_r))[0]
	
	x = np.array(zetas)[indices]
	y = np.array(xis)[indices]
	z = np.array(entropies)[indices] / np.max(np.array(entropies)[indices])
	c = np.array(princ_coords)[indices]
	Tools.Plot_3DShape(x, y, z, c, r'$\zeta$ [\AA]', r'$\xi$ [\AA]', r'$\mathcal{S}$ (norm.)', r'$\perp$-coord.')

	x = np.array(zetas)[indices] / np.array(xis)[indices]  
	y = np.array(sizes)[indices]
	z = np.array(entropies)[indices] / np.max(np.array(entropies)[indices])
	c = np.array(princ_coords)[indices]
	Tools.Plot_3DShape(x, y, z, c, r'$\nu$', r'$s$ [\AA]', r'$\mathcal{S}$ (norm.)', r'$\perp$-coord.')

	x = np.array(l_is)[indices] - np.array(pore_radii)[indices] 
	y = np.array(zetas)[indices] / np.array(xis)[indices]     
	z = np.array(princ_coords)[indices] 
	c = np.array(entropies)[indices] / np.max(np.array(entropies)[indices])
	Tools.Plot_3DShape(x, y, z, c, r'$l_{i}$ [\AA]', r'$\nu$', r'$\perp$-coord.', r'$\mathcal{S}$ (norm.)', location = "left")

	x = np.array(l_is)[indices] - np.array(pore_radii)[indices] 
	y = np.array(zetas)[indices] / np.array(xis)[indices]     
	z = np.array(princ_coords)[indices] 
	c = np.array(entropies)[indices] / np.max(np.array(entropies)[indices])
	Tools.Plot_3DShape(x, y, z, c, r'$l_{i}$ [\AA]', r'$\nu$', r'$\perp$-coord.', r'$\mathcal{S}$ (norm.)', location = "left")


	'''	
			Cumulative atom number
	'''
	# Prepare the averages
	ns_ = np.zeros( (int(len(m0s) / ModelParameters.N_SCALES), ModelParameters.N_SCALES) )
	n_models_ = np.zeros( (int(len(m0s) / ModelParameters.N_SCALES), ModelParameters.N_SCALES) )
	
	for i in range(int(len(m0s) / ModelParameters.N_SCALES)):
		ind_s = i*ModelParameters.N_SCALES 
		ind_e = (i+1)*ModelParameters.N_SCALES 
		ns_[i,] = ns[ind_s:ind_e]
		n_models_[i,] = n_models[ind_s:ind_e]

	Tools.Plot_CumulativeAtomNumber(ns_, n_models_, min(ind_l_is), max(ind_l_is), [np.array(MAFEs), np.array(pvals), np.array(fracDims_i), np.array(fracDims_i_exper), np.array(fracDims_max)], statsType = 'mean_of_medians')

	'''
			Hydropathic energy per atom
	'''
	# Prepare averages
	m0s_ = np.zeros( (int(len(m0s) / ModelParameters.N_SCALES), ModelParameters.N_SCALES) )
	m0s_pho_ = np.zeros( (int(len(m0s) / ModelParameters.N_SCALES), ModelParameters.N_SCALES) )
	m0s_phi_ = np.zeros( (int(len(m0s) / ModelParameters.N_SCALES),ModelParameters.N_SCALES) )
	m0_models_ = np.zeros( (int(len(m0s) / ModelParameters.N_SCALES),ModelParameters.N_SCALES) )
	
	for i in range(int(len(m0s) / ModelParameters.N_SCALES)):
		ind_s = i*ModelParameters.N_SCALES 
		ind_e = (i+1)*ModelParameters.N_SCALES 
		m0s_[i,] = m0s[ind_s:ind_e]
		m0s_pho_[i,] = m0s_pho[ind_s:ind_e]
		m0s_phi_[i,] = m0s_phi[ind_s:ind_e]
		m0_models_[i,] = m0_models[ind_s:ind_e]

	Tools.Plot_AtomicHydropathicEnergy(m0s_, m0s_pho_, m0s_phi_, m0_models_, min(ind_l_is), max(ind_l_is), [xis, zetas], statsType = 'mean_of_medians')

	'''
			Dipole moment (first-order hydr mom)
	'''
	# Prepare averages
	# Prepare averages
	hz_poss_ = np.zeros( (int(len(m0s) / ModelParameters.N_SCALES) ,ModelParameters.N_SCALES) )
	hz_ESs_ = np.zeros( (int(len(m0s) / ModelParameters.N_SCALES) ,ModelParameters.N_SCALES) )
	hz_negs_ = np.zeros( (int(len(m0s) / ModelParameters.N_SCALES) ,ModelParameters.N_SCALES) )
	hz_ISs_ = np.zeros( (int(len(m0s) / ModelParameters.N_SCALES) ,ModelParameters.N_SCALES) )

	for i in range(int(len(m0s) / ModelParameters.N_SCALES)):
		
		ind_s = i*ModelParameters.N_SCALES 
		ind_e = (i+1)*ModelParameters.N_SCALES 

		hz_poss_[i,] = data_h_z_poss[ind_s:ind_e]
		hz_ESs_[i,] = data_h_z_ESs[ind_s:ind_e]
		hz_negs_[i,] = data_h_z_negs[ind_s:ind_e]
		hz_ISs_[i,] = data_h_z_ISs[ind_s:ind_e]

	Tools.Plot_DipoleMoment(hz_poss_, hz_ISs_, hz_negs_, hz_ESs_, min(ind_l_is), max(ind_l_is), [expPres, expInfl2s], [PC_expPres, PC_expInfl2s], statsType = 'mean_of_medians', orderOfMom = orderOfMom, x_min = -30, x_max = 30)

